# -*- coding: utf-8 -*-





import os


import uuid

import hashlib


import re


import xml.etree.ElementTree as ET

from datetime import datetime



from qgis.core import (

    QgsRasterLayer,

    QgsVectorLayer,

    QgsProject,

    QgsCoordinateReferenceSystem,

    QgsCoordinateTransform,

    QgsMapLayer,

)



from qgis.PyQt.QtCore import Qt, QSettings


from qgis.PyQt.QtWidgets import (

    QDialog,

    QVBoxLayout,

    QHBoxLayout,

    QFormLayout,

    QLabel,

    QLineEdit,

    QTextEdit,

    QPushButton,

    QFileDialog,

    QMessageBox,

    QScrollArea,

    QWidget,

    QTabWidget,

    QComboBox,

    QCheckBox,

    QGroupBox,

)





SOI_TEMPLATE_FIELDS = [

    {

        "id": "f_001",

        "section": "Dataset File Information",

        "label": "File name",

        "iso_path": "(Dataset file name — not an ISO property)",

        "type": "Free Text",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_002",

        "section": "Dataset File Information",

        "label": "Brief description",

        "iso_path": "(Short dataset description — not an ISO property)",

        "type": "Free Text",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_003",

        "section": "MD_Metadata › Identifier",

        "label": "metadataIdentifier · code",

        "iso_path": "MD_Metadata . metadataIdentifier . MD_Identifier . code",

        "type": "Free Text",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_004",

        "section": "MD_Metadata › Identifier",

        "label": "metadataIdentifier · description",

        "iso_path": "MD_Metadata . metadataIdentifier . MD_Identifier . description",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_005",

        "section": "MD_Metadata › Date & Standard",

        "label": "dateInfo · date",

        "iso_path": "MD_Metadata . dateInfo . CI_Date . date",

        "type": "DateTime",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_006",

        "section": "MD_Metadata › Date & Standard",

        "label": "dateInfo · dateType",

        "iso_path": "MD_Metadata . dateInfo . CI_Date . dateType (CI_DateTypeCode)",

        "type": "CodeList",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_007",

        "section": "MD_Metadata › Date & Standard",

        "label": "metadataStandard · title",

        "iso_path": "MD_Metadata . metadataStandard . CI_Citation . title",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_008",

        "section": "MD_Metadata › Date & Standard",

        "label": "metadataStandard · edition",

        "iso_path": "MD_Metadata . metadataStandard . CI_Citation . edition",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_009",

        "section": "MD_Metadata › Date & Standard",

        "label": "metadataStandard · editionDate",

        "iso_path": "MD_Metadata . metadataStandard . CI_Citation . date[dateType=edition] . CI_Date . date",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_010",

        "section": "MD_Metadata › Contact",

        "label": "contact · role",

        "iso_path": "MD_Metadata . contact . CI_Responsibility . role (CI_RoleCode)",

        "type": "CodeList",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_011",

        "section": "MD_Metadata › Contact",

        "label": "contact · party name",

        "iso_path": "MD_Metadata . contact . CI_Responsibility . party . CI_Organisation . name",

        "type": "Free Text",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_012",

        "section": "MD_Metadata › Contact",

        "label": "contact · electronicMailAddress",

        "iso_path": "MD_Metadata . contact . CI_Responsibility . party . CI_Organisation . contactInfo . CI_Contact . address . CI_Address . electronicMailAddress",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_013",

        "section": "MD_Metadata › Contact",

        "label": "contact · onlineResource linkage",

        "iso_path": "MD_Metadata . contact . CI_Responsibility . party . CI_Organisation . contactInfo . CI_Contact . onlineResource . CI_OnlineResource . linkage",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_014",

        "section": "MD_Metadata › Scope",

        "label": "metadataScope · resourceScope",

        "iso_path": "MD_Metadata . metadataScope . MD_MetadataScope . resourceScope (MD_ScopeCode)",

        "type": "CodeList",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_015",

        "section": "MD_DataIdentification › Citation",

        "label": "citation · title",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . citation . CI_Citation . title",

        "type": "Free Text",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_016",

        "section": "MD_DataIdentification › Citation",

        "label": "citation · alternateTitle",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . citation . CI_Citation . alternateTitle",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_017",

        "section": "MD_DataIdentification › Citation",

        "label": "citation · date",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . citation . CI_Citation . date . CI_Date . date",

        "type": "DateTime",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_018",

        "section": "MD_DataIdentification › Citation",

        "label": "citation · dateType",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . citation . CI_Citation . date . CI_Date . dateType (CI_DateTypeCode)",

        "type": "CodeList",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_019",

        "section": "MD_DataIdentification › Citation",

        "label": "citation · identifier · code",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . citation . CI_Citation . identifier . MD_Identifier . code",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_020",

        "section": "MD_DataIdentification › Citation",

        "label": "citation · edition",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . citation . CI_Citation . edition",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_021",

        "section": "MD_DataIdentification › Citation",

        "label": "citation · editionDate",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . citation . CI_Citation . date[dateType=edition] . CI_Date . date",

        "type": "DateTime",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_022",

        "section": "MD_DataIdentification › Citation",

        "label": "citation · presentationForm",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . citation . CI_Citation . presentationForm (CI_PresentationFormCode)",

        "type": "CodeList",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_023",

        "section": "MD_DataIdentification › Citation",

        "label": "citation · series · name",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . citation . CI_Citation . series . CI_Series . name",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_024",

        "section": "MD_DataIdentification › General",

        "label": "abstract",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . abstract",

        "type": "Free Text",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_025",

        "section": "MD_DataIdentification › General",

        "label": "purpose",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . purpose",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_026",

        "section": "MD_DataIdentification › General",

        "label": "credit",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . credit",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_027",

        "section": "MD_DataIdentification › General",

        "label": "spatialRepresentationType",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . spatialRepresentationType (MD_SpatialRepresentationTypeCode)",

        "type": "CodeList",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_028",

        "section": "MD_DataIdentification › General",

        "label": "spatialResolution · equivalentScale",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . spatialResolution . MD_Resolution . equivalentScale . MD_RepresentativeFraction . denominator",

        "type": "Numeric",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_029",

        "section": "MD_DataIdentification › General",

        "label": "spatialResolution · distance (GSD)",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . spatialResolution . MD_Resolution . distance",

        "type": "Numeric",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_030",

        "section": "MD_DataIdentification › General",

        "label": "levelOfDetail",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . levelOfDetail",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_031",

        "section": "MD_DataIdentification › General",

        "label": "temporalResolution",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . temporalResolution",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_032",

        "section": "MD_DataIdentification › General",

        "label": "topicCategory",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . topicCategory (MD_TopicCategoryCode)",

        "type": "CodeList",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_033",

        "section": "MD_DataIdentification › General",

        "label": "graphicOverview · fileName",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . graphicOverview . MD_BrowseGraphic . fileName",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_034",

        "section": "MD_DataIdentification › General",

        "label": "graphicOverview · fileDescription",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . graphicOverview . MD_BrowseGraphic . fileDescription",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_035",

        "section": "CI_Responsibility › Point of Contact",

        "label": "pointOfContact · role",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . pointOfContact . CI_Responsibility . role (CI_RoleCode)",

        "type": "CodeList",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_036",

        "section": "CI_Responsibility › Point of Contact",

        "label": "pointOfContact · party · organisation name",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . pointOfContact . CI_Responsibility . party . CI_Organisation . name",

        "type": "Free Text",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_037",

        "section": "CI_Responsibility › Point of Contact",

        "label": "pointOfContact · party · individual name",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . pointOfContact . CI_Responsibility . party . CI_Individual . name",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_038",

        "section": "CI_Responsibility › Point of Contact",

        "label": "pointOfContact · party · positionName",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . pointOfContact . CI_Responsibility . party . CI_Individual . positionName",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_039",

        "section": "CI_Responsibility › Point of Contact",

        "label": "pointOfContact · electronicMailAddress",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . pointOfContact . CI_Responsibility . party . CI_Organisation . contactInfo . CI_Contact . address . CI_Address . electronicMailAddress",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_040",

        "section": "CI_Responsibility › Point of Contact",

        "label": "pointOfContact · voice",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . pointOfContact . CI_Responsibility . party . CI_Organisation . contactInfo . CI_Contact . phone . CI_Telephone . voice",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_041",

        "section": "CI_Responsibility › Point of Contact",

        "label": "pointOfContact · onlineResource linkage",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . pointOfContact . CI_Responsibility . party . CI_Organisation . contactInfo . CI_Contact . onlineResource . CI_OnlineResource . linkage",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_042",

        "section": "MD_Keywords",

        "label": "keyword · theme",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . descriptiveKeywords . MD_Keywords . keyword [type=theme]",

        "type": "Free Text",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_043",

        "section": "MD_Keywords",

        "label": "keyword · place",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . descriptiveKeywords . MD_Keywords . keyword [type=place]",

        "type": "Free Text",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_044",

        "section": "MD_Keywords",

        "label": "keyword · discipline",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . descriptiveKeywords . MD_Keywords . keyword [type=discipline]",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_045",

        "section": "MD_Keywords",

        "label": "keyword · stratum",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . descriptiveKeywords . MD_Keywords . keyword [type=stratum]",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_046",

        "section": "MD_Keywords",

        "label": "keyword · temporal",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . descriptiveKeywords . MD_Keywords . keyword [type=temporal]",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_047",

        "section": "MD_Keywords",

        "label": "keyword · product",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . descriptiveKeywords . MD_Keywords . keyword [type=theme] (product)",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_048",

        "section": "MD_Keywords",

        "label": "keyword · platform",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . descriptiveKeywords . MD_Keywords . keyword [type=theme] (platform)",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_049",

        "section": "MD_Keywords",

        "label": "thesaurusName · title",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . descriptiveKeywords . MD_Keywords . thesaurusName . CI_Citation . title",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_050",

        "section": "MD_Keywords",

        "label": "thesaurusName · date",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . descriptiveKeywords . MD_Keywords . thesaurusName . CI_Citation . date . CI_Date . date",

        "type": "DateTime",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_051",

        "section": "MD_SpatialRepresentation",

        "label": "topologyLevel",

        "iso_path": "MD_Metadata . spatialRepresentationInfo . MD_VectorSpatialRepresentation . topologyLevel (MD_TopologyLevelCode)",

        "type": "CodeList",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_052",

        "section": "MD_SpatialRepresentation",

        "label": "geometricObjects · geometricObjectType",

        "iso_path": "MD_Metadata . spatialRepresentationInfo . MD_VectorSpatialRepresentation . geometricObjects . MD_GeometricObjects . geometricObjectType",

        "type": "CodeList",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_053",

        "section": "MD_SpatialRepresentation",

        "label": "geometricObjects · geometricObjectCount",

        "iso_path": "MD_Metadata . spatialRepresentationInfo . MD_VectorSpatialRepresentation . geometricObjects . MD_GeometricObjects . geometricObjectCount",

        "type": "Numeric",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_054",

        "section": "MD_SpatialRepresentation",

        "label": "numberOfDimensions",

        "iso_path": "MD_Metadata . spatialRepresentationInfo . MD_GridSpatialRepresentation . numberOfDimensions",

        "type": "Numeric",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_055",

        "section": "MD_SpatialRepresentation",

        "label": "axisDimensionProperties · column · dimensionSize",

        "iso_path": "MD_Metadata . spatialRepresentationInfo . MD_GridSpatialRepresentation . axisDimensionProperties . MD_Dimension [column] . dimensionSize",

        "type": "Numeric",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_056",

        "section": "MD_SpatialRepresentation",

        "label": "axisDimensionProperties · row · dimensionSize",

        "iso_path": "MD_Metadata . spatialRepresentationInfo . MD_GridSpatialRepresentation . axisDimensionProperties . MD_Dimension [row] . dimensionSize",

        "type": "Numeric",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_057",

        "section": "MD_SpatialRepresentation",

        "label": "cellGeometry",

        "iso_path": "MD_Metadata . spatialRepresentationInfo . MD_GridSpatialRepresentation . cellGeometry (MD_CellGeometryCode)",

        "type": "CodeList",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_058",

        "section": "MD_SpatialRepresentation",

        "label": "transformationParameterAvailability",

        "iso_path": "MD_Metadata . spatialRepresentationInfo . MD_GridSpatialRepresentation . transformationParameterAvailability",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_059",

        "section": "EX_Extent",

        "label": "extent · description",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . extent . EX_Extent . description",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_060",

        "section": "EX_Extent",

        "label": "geographicElement · westBoundLongitude",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . extent . EX_Extent . geographicElement . EX_GeographicBoundingBox . westBoundLongitude",

        "type": "Numeric",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_061",

        "section": "EX_Extent",

        "label": "geographicElement · eastBoundLongitude",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . extent . EX_Extent . geographicElement . EX_GeographicBoundingBox . eastBoundLongitude",

        "type": "Numeric",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_062",

        "section": "EX_Extent",

        "label": "geographicElement · southBoundLatitude",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . extent . EX_Extent . geographicElement . EX_GeographicBoundingBox . southBoundLatitude",

        "type": "Numeric",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_063",

        "section": "EX_Extent",

        "label": "geographicElement · northBoundLatitude",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . extent . EX_Extent . geographicElement . EX_GeographicBoundingBox . northBoundLatitude",

        "type": "Numeric",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_064",

        "section": "EX_Extent",

        "label": "verticalElement · minimumValue",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . extent . EX_Extent . verticalElement . EX_VerticalExtent . minimumValue",

        "type": "Numeric",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_065",

        "section": "EX_Extent",

        "label": "verticalElement · maximumValue",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . extent . EX_Extent . verticalElement . EX_VerticalExtent . maximumValue",

        "type": "Numeric",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_066",

        "section": "EX_Extent",

        "label": "temporalElement · extent",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . extent . EX_Extent . temporalElement . EX_TemporalExtent . extent",

        "type": "DateTime",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_067",

        "section": "RS_ReferenceSystem › CRS",

        "label": "Horizontal CRS · code",

        "iso_path": "MD_Metadata . referenceSystemInfo . MD_ReferenceSystem . referenceSystemIdentifier . RS_Identifier . code (Horizontal)",

        "type": "Free Text",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_068",

        "section": "RS_ReferenceSystem › CRS",

        "label": "Vertical CRS · code",

        "iso_path": "MD_Metadata . referenceSystemInfo . MD_ReferenceSystem . referenceSystemIdentifier . RS_Identifier . code (Vertical)",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_069",

        "section": "RS_ReferenceSystem › CRS",

        "label": "Vertical datum description",

        "iso_path": "MD_Metadata . referenceSystemInfo . MD_ReferenceSystem . referenceSystemIdentifier . RS_Identifier . description",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_070",

        "section": "RS_ReferenceSystem › CRS",

        "label": "CRS description / notes",

        "iso_path": "MD_Metadata . referenceSystemInfo . MD_ReferenceSystem (notes)",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_071",

        "section": "DQ_DataQuality",

        "label": "scope · level",

        "iso_path": "MD_Metadata . dataQualityInfo . DQ_DataQuality . scope . MD_Scope . level (MD_ScopeCode)",

        "type": "CodeList",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_072",

        "section": "DQ_DataQuality",

        "label": "report · DQ_AbsoluteExternalPositionalAccuracy · result · statement",

        "iso_path": "MD_Metadata . dataQualityInfo . DQ_DataQuality . report . DQ_AbsoluteExternalPositionalAccuracy . result . DQ_DescriptiveResult . statement",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_073",

        "section": "DQ_DataQuality",

        "label": "report · DQ_AbsoluteExternalPositionalAccuracy (vertical) · result · statement",

        "iso_path": "MD_Metadata . dataQualityInfo . DQ_DataQuality . report . DQ_AbsoluteExternalPositionalAccuracy [vertical] . result . DQ_DescriptiveResult . statement",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_074",

        "section": "DQ_DataQuality",

        "label": "report · DQ_CompletenessOmission · result · statement",

        "iso_path": "MD_Metadata . dataQualityInfo . DQ_DataQuality . report . DQ_CompletenessOmission . result . DQ_DescriptiveResult . statement",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_075",

        "section": "DQ_DataQuality",

        "label": "report · DQ_TopologicalConsistency · result · statement",

        "iso_path": "MD_Metadata . dataQualityInfo . DQ_DataQuality . report . DQ_TopologicalConsistency . result . DQ_DescriptiveResult . statement",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_076",

        "section": "DQ_DataQuality",

        "label": "report · DQ_ConceptualConsistency · result · statement",

        "iso_path": "MD_Metadata . dataQualityInfo . DQ_DataQuality . report . DQ_ConceptualConsistency . result . DQ_DescriptiveResult . statement",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_077",

        "section": "DQ_DataQuality",

        "label": "report · DQ_FormatConsistency · result · statement",

        "iso_path": "MD_Metadata . dataQualityInfo . DQ_DataQuality . report . DQ_FormatConsistency . result . DQ_DescriptiveResult . statement",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_078",

        "section": "DQ_DataQuality",

        "label": "report · DQ_TemporalConsistency · result · statement",

        "iso_path": "MD_Metadata . dataQualityInfo . DQ_DataQuality . report . DQ_TemporalConsistency . result . DQ_DescriptiveResult . statement",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_079",

        "section": "DQ_DataQuality",

        "label": "report · DQ_ThematicClassificationCorrectness · result · statement",

        "iso_path": "MD_Metadata . dataQualityInfo . DQ_DataQuality . report . DQ_ThematicClassificationCorrectness . result . DQ_DescriptiveResult . statement",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_080",

        "section": "DQ_DataQuality",

        "label": "report · DQ_QuantitativeAttributeAccuracy · result · statement",

        "iso_path": "MD_Metadata . dataQualityInfo . DQ_DataQuality . report . DQ_QuantitativeAttributeAccuracy . result . DQ_DescriptiveResult . statement",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_081",

        "section": "MD_Constraints",

        "label": "accessConstraints",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . resourceConstraints . MD_LegalConstraints . accessConstraints (MD_RestrictionCode)",

        "type": "CodeList",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_082",

        "section": "MD_Constraints",

        "label": "useConstraints",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . resourceConstraints . MD_LegalConstraints . useConstraints (MD_RestrictionCode)",

        "type": "CodeList",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_083",

        "section": "MD_Constraints",

        "label": "otherConstraints",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . resourceConstraints . MD_LegalConstraints . otherConstraints",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_084",

        "section": "MD_Constraints",

        "label": "classification",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . resourceConstraints . MD_SecurityConstraints . classification (MD_ClassificationCode)",

        "type": "CodeList",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_085",

        "section": "MD_Constraints",

        "label": "MD_SecurityConstraints · userNote",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . resourceConstraints . MD_SecurityConstraints . userNote",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_086",

        "section": "MD_Constraints",

        "label": "useLimitation",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . resourceConstraints . MD_Constraints . useLimitation",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_087",

        "section": "MD_ContentInformation",

        "label": "contentType",

        "iso_path": "MD_Metadata . contentInfo . MD_CoverageDescription . contentType (MD_CoverageContentTypeCode)",

        "type": "CodeList",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_088",

        "section": "MD_ContentInformation",

        "label": "attributeDescription",

        "iso_path": "MD_Metadata . contentInfo . MD_CoverageDescription . attributeDescription",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_089",

        "section": "FC_FeatureCatalogue",

        "label": "featureCatalogue · title",

        "iso_path": "MD_Metadata . contentInfo . MD_FeatureCatalogueDescription . featureCatalogueCitation . CI_Citation . title",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_090",

        "section": "FC_FeatureCatalogue",

        "label": "featureCatalogue · date",

        "iso_path": "MD_Metadata . contentInfo . MD_FeatureCatalogueDescription . featureCatalogueCitation . CI_Citation . date . CI_Date . date",

        "type": "DateTime",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_091",

        "section": "FC_FeatureCatalogue",

        "label": "featureCatalogue · edition",

        "iso_path": "MD_Metadata . contentInfo . MD_FeatureCatalogueDescription . featureCatalogueCitation . CI_Citation . edition",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_092",

        "section": "FC_FeatureCatalogue",

        "label": "includedWithDataset",

        "iso_path": "MD_Metadata . contentInfo . MD_FeatureCatalogueDescription . includedWithDataset",

        "type": "CodeList",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_093",

        "section": "FC_FeatureCatalogue",

        "label": "featureCatalogue · responsible party",

        "iso_path": "MD_Metadata . contentInfo . MD_FeatureCatalogueDescription . featureCatalogueCitation . CI_Citation . citedResponsibleParty . CI_Responsibility . party . CI_Organisation . name",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_094",

        "section": "MD_Distribution",

        "label": "distributionFormat · name",

        "iso_path": "MD_Metadata . distributionInfo . MD_Distribution . distributionFormat . MD_Format . name",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_095",

        "section": "MD_Distribution",

        "label": "distributionFormat · version",

        "iso_path": "MD_Metadata . distributionInfo . MD_Distribution . distributionFormat . MD_Format . version",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_096",

        "section": "MD_Distribution",

        "label": "distributor · organisation",

        "iso_path": "MD_Metadata . distributionInfo . MD_Distribution . distributor . MD_Distributor . distributorContact . CI_Responsibility . party . CI_Organisation . name",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_097",

        "section": "MD_Distribution",

        "label": "distributor · electronicMailAddress",

        "iso_path": "MD_Metadata . distributionInfo . MD_Distribution . distributor . MD_Distributor . distributorContact . CI_Responsibility . party . CI_Organisation . contactInfo . CI_Contact . address . CI_Address . electronicMailAddress",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_098",

        "section": "MD_Distribution",

        "label": "transferOptions · offLine · mediumName",

        "iso_path": "MD_Metadata . distributionInfo . MD_Distribution . transferOptions . MD_DigitalTransferOptions . offLine . MD_Medium . name",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_099",

        "section": "LI_Lineage & Maintenance",

        "label": "resourceLineage · statement",

        "iso_path": "MD_Metadata . resourceLineage . LI_Lineage . statement",

        "type": "Free Text",

        "required": "Yes",

        "default": ""

    },

    {

        "id": "f_100",

        "section": "LI_Lineage & Maintenance",

        "label": "metadataMaintenance · maintenanceAndUpdateFrequency",

        "iso_path": "MD_Metadata . metadataMaintenance . MD_MaintenanceInformation . maintenanceAndUpdateFrequency (MD_MaintenanceFrequencyCode)",

        "type": "CodeList",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_101",

        "section": "LI_Lineage & Maintenance",

        "label": "metadataMaintenance · maintenanceNote",

        "iso_path": "MD_Metadata . metadataMaintenance . MD_MaintenanceInformation . maintenanceNote",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_102",

        "section": "Advanced / 3D",

        "label": "MD_Dimension · dimension (2D or 3D)",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . spatialRepresentationInfo . MD_VectorSpatialRepresentation . topologyLevel / MD_Dimension",

        "type": "Numeric",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_103",

        "section": "Advanced / 3D",

        "label": "3D CRS / additional CRS notes",

        "iso_path": "MD_ReferenceSystem (for 3D datasets — combined H+V CRS)",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_104",

        "section": "Advanced / 3D",

        "label": "Point cloud density / classification",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . supplementalInformation (LiDAR details)",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_105",

        "section": "Advanced / 3D",

        "label": "DEM / Geoid notes",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . supplementalInformation (DEM/Geoid specifics)",

        "type": "Free Text",

        "required": "No",

        "default": ""

    },

    {

        "id": "f_106",

        "section": "Advanced / 3D",

        "label": "supplementalInformation",

        "iso_path": "MD_Metadata . identificationInfo . MD_DataIdentification . supplementalInformation",

        "type": "Free Text",

        "required": "No",

        "default": ""

    }

]





# These fields are normally derived from the dataset and should not be manually edited.

# They are shown in grey in the GUI.

AUTO_DERIVED_LABELS = {

    "File name",

    "geographicElement · westBoundLongitude",

    "geographicElement · eastBoundLongitude",

    "geographicElement · southBoundLatitude",

    "geographicElement · northBoundLatitude",

    "Horizontal CRS · code",

    "CRS description / notes",

    "axisDimensionProperties · column · dimensionSize",

    "axisDimensionProperties · row · dimensionSize",

    "numberOfDimensions",

    "cellGeometry",

    "spatialResolution · distance (GSD)",

    "MD_Dimension · dimension (2D or 3D)",

}



MULTILINE_HINTS = (

    "abstract",

    "description",

    "statement",

    "purpose",

    "credit",

    "constraints",

    "note",

    "lineage",

    "supplemental",

    "accuracy",

    "consistency",

    "limitation",

)



CODELIST_VALUES = {

    "dateInfo · dateType": ["creation", "publication", "revision"],

    "citation · dateType": ["creation", "publication", "revision"],

    "contact · role": ["pointOfContact", "custodian", "owner", "distributor", "originator", "processor", "publisher", "author"],

    "metadataScope · resourceScope": ["dataset", "series", "service", "model", "tile"],

    "spatialRepresentationType": ["grid", "vector", "tin", "textTable", "stereoModel", "video"],

    "topicCategory": ["imageryBaseMapsEarthCover", "planningCadastre", "boundaries", "elevation", "geoscientificInformation", "location", "structure"],

    "pointOfContact · role": ["pointOfContact", "custodian", "owner", "distributor", "originator", "processor", "publisher", "author"],

    "topologyLevel": ["geometryOnly", "topology1D", "planarGraph", "fullPlanarGraph", "surfaceGraph", "fullSurfaceGraph", "topology3D"],

    "geometricObjects · geometricObjectType": ["complex", "composite", "curve", "point", "solid", "surface"],

    "cellGeometry": ["area", "point"],

    "scope · level": ["dataset", "series", "tile", "model"],

    "accessConstraints": ["copyright", "license", "intellectualPropertyRights", "restricted", "otherRestrictions"],

    "useConstraints": ["copyright", "license", "intellectualPropertyRights", "restricted", "otherRestrictions"],

    "classification": ["unclassified", "restricted", "confidential", "secret", "topSecret"],

    "contentType": ["image", "thematicClassification", "physicalMeasurement"],

    "includedWithDataset": ["true", "false"],

    "metadataMaintenance · maintenanceAndUpdateFrequency": ["continual", "daily", "weekly", "monthly", "annually", "asNeeded", "notPlanned", "unknown"],

}





def safe_text(value):

    if value is None:

        return ""

    return str(value).strip()





def xml_safe_tag(text):

    tag = safe_text(text)

    for ch in [" ", "·", "/", "\\", "-", "›", "(", ")", ".", ",", "&", ":", ";"]:

        tag = tag.replace(ch, "_")

    while "__" in tag:

        tag = tag.replace("__", "_")

    tag = tag.strip("_")

    if not tag:

        tag = "field"

    if tag[0].isdigit():

        tag = "field_" + tag

    return tag





def today_iso():

    return datetime.now().strftime("%Y-%m-%d")





def now_iso():

    return datetime.now().strftime("%Y-%m-%dT%H:%M:%S")





def extract_year(value):

    m = re.search(r"(19|20)\d{2}", safe_text(value))

    return m.group(0) if m else ""





def is_multiline_field(label, iso_path):

    s = (safe_text(label) + " " + safe_text(iso_path)).lower()

    return any(h in s for h in MULTILINE_HINTS)





def list_ogr_sublayers(container_path):

    """List vector layers inside GeoPackage / FileGDB using GDAL/OGR when available."""

    layers = []

    try:

        from osgeo import ogr

        ds = ogr.Open(container_path)

        if ds:

            for i in range(ds.GetLayerCount()):

                lyr = ds.GetLayerByIndex(i)

                if lyr:

                    layers.append(lyr.GetName())

    except Exception:

        pass

    return layers





def list_gdal_raster_subdatasets(container_path):

    """List raster datasets inside FileGDB/GeoPackage using GDAL subdatasets.



    FileGDB raster datasets are not normal files. GDAL exposes them as

    subdataset URIs such as OpenFileGDB:"C:/path/data.gdb":RasterName.

    QGIS must open that returned URI, not the .gdb folder path.

    """

    rasters = []

    try:

        from osgeo import gdal



        # Try raster-only first.

        ds = gdal.OpenEx(container_path, gdal.OF_RASTER)

        if ds is None:

            ds = gdal.OpenEx(container_path)

        if ds is None:

            return rasters



        subdatasets = ds.GetSubDatasets() or []

        seen = set()

        for uri, desc in subdatasets:

            display = desc or uri

            # Make the combo text readable. Common desc examples include:

            #   OpenFileGDB:"C:/x.gdb":RasterName - RasterName

            # Keep the URI as itemData because it is what QgsRasterLayer needs.

            if uri not in seen:

                rasters.append({"name": display, "uri": uri, "type": "raster"})

                seen.add(uri)



        # Some containers expose a single raster directly without subdatasets.

        if not rasters and getattr(ds, "RasterCount", 0) > 0:

            rasters.append({

                "name": os.path.basename(container_path),

                "uri": container_path,

                "type": "raster",

            })

    except Exception:

        pass

    return rasters





def list_container_datasets(container_path):

    """Return both vector layers and raster subdatasets from .gdb/.gpkg."""

    datasets = []



    for name in list_ogr_sublayers(container_path):

        datasets.append({

            "name": name,

            "uri": f"{container_path}|layername={name}",

            "type": "vector",

        })



    for raster in list_gdal_raster_subdatasets(container_path):

        datasets.append(raster)



    return datasets



def extract_layer_metadata(layer, source_path=""):

    """Extract common metadata from a QGIS raster/vector layer.



    This is better than file-only extraction because it also works for FileGDB,

    GeoPackage, PostGIS, WMS/WFS and any layer already loaded in QGIS.

    """

    result = {

        "valid": False,

        "file_path": source_path or "",

        "File name": "",

        "file_size_mb": "",

        "provider": "",

        "layer_type": "",

        "width": "",

        "height": "",

        "bands": "",

        "feature_count": "",

        "geometry_type": "",

        "crs_authid": "",

        "crs_description": "",

        "extent_layer_crs": "",

        "west": "",

        "east": "",

        "south": "",

        "north": "",

        "pixel_width": "",

        "pixel_height": "",

        "number_of_dimensions": "2",

        "cell_geometry": "area",

        "format_name": "",

    }



    if layer is None or not layer.isValid():

        return result



    result["valid"] = True

    result["File name"] = layer.name()

    result["provider"] = layer.providerType()



    # Try to get the real source path from QGIS.

    try:

        result["file_path"] = source_path or layer.source()

    except Exception:

        pass



    path_for_size = result["file_path"].split("|")[0]

    if path_for_size and os.path.exists(path_for_size) and os.path.isfile(path_for_size):

        result["file_size_mb"] = round(os.path.getsize(path_for_size) / (1024 * 1024), 3)



    try:

        if layer.type() == QgsMapLayer.RasterLayer:

            result["layer_type"] = "Raster"

            result["width"] = layer.width()

            result["height"] = layer.height()

            result["bands"] = layer.bandCount()

            if layer.width() > 0:

                result["pixel_width"] = abs(layer.extent().width() / layer.width())

            if layer.height() > 0:

                result["pixel_height"] = abs(layer.extent().height() / layer.height())

            result["format_name"] = os.path.splitext(path_for_size)[1].replace(".", "").upper() or result["provider"]

        elif layer.type() == QgsMapLayer.VectorLayer:

            result["layer_type"] = "Vector"

            result["feature_count"] = layer.featureCount()

            try:

                result["geometry_type"] = layer.geometryType()

            except Exception:

                result["geometry_type"] = ""

            result["format_name"] = os.path.splitext(path_for_size)[1].replace(".", "").upper() or result["provider"]

    except Exception:

        pass



    crs = layer.crs()

    result["crs_authid"] = crs.authid()

    result["crs_description"] = crs.description()



    extent = layer.extent()

    result["extent_layer_crs"] = (

        f"xMin={extent.xMinimum()}, yMin={extent.yMinimum()}, "

        f"xMax={extent.xMaximum()}, yMax={extent.yMaximum()}"

    )



    try:

        dest_crs = QgsCoordinateReferenceSystem("EPSG:4326")

        transform = QgsCoordinateTransform(crs, dest_crs, QgsProject.instance())

        geo_extent = transform.transformBoundingBox(extent)

        result["west"] = geo_extent.xMinimum()

        result["east"] = geo_extent.xMaximum()

        result["south"] = geo_extent.yMinimum()

        result["north"] = geo_extent.yMaximum()

    except Exception:

        result["west"] = extent.xMinimum()

        result["east"] = extent.xMaximum()

        result["south"] = extent.yMinimum()

        result["north"] = extent.yMaximum()



    return result





def write_metadata_xml(values, output_xml_path):

    """Write ISO 19115-1 / ISO 19115-3 style XML using proper namespaces.



    This replaces the older custom SOI XML writer. It keeps your existing UI field IDs

    and maps them into ISO namespace elements such as mdb:MD_Metadata,

    mri:MD_DataIdentification, cit:CI_Citation, gex:EX_GeographicBoundingBox, etc.

    """

    NS = {

        "gco": "https://schemas.isotc211.org/19103/-/gco/1.2",

        "gcx": "https://schemas.isotc211.org/19103/-/gcx/1.2",

        "cit": "https://schemas.isotc211.org/19115/-1/cit/1.3",

        "gex": "https://schemas.isotc211.org/19115/-1/gex/1.3",

        "lan": "https://schemas.isotc211.org/19115/-1/lan/1.3",

        "mcc": "https://schemas.isotc211.org/19115/-1/mcc/1.3",

        "mdb": "https://schemas.isotc211.org/19115/-1/mdb/1.3",

        "mri": "https://schemas.isotc211.org/19115/-1/mri/1.3",

        "mrs": "https://schemas.isotc211.org/19115/-1/mrs/1.3",

        "mrd": "https://schemas.isotc211.org/19115/-1/mrd/1.3",

        "mrl": "https://schemas.isotc211.org/19115/-1/mrl/1.3",

        "mco": "https://schemas.isotc211.org/19115/-1/mco/1.3",

        "msr": "https://schemas.isotc211.org/19115/-1/msr/1.3",

        "xsi": "http://www.w3.org/2001/XMLSchema-instance",

        "gml": "http://www.opengis.net/gml/3.2",

        "xlink": "http://www.w3.org/1999/xlink",

    }

    for prefix, uri in NS.items():

        ET.register_namespace(prefix, uri)



    def q(prefix, tag):

        return f"{{{NS[prefix]}}}{tag}"



    def get(label, default=""):

        fid = next((f["id"] for f in SOI_TEMPLATE_FIELDS if f["label"] == label), None)

        return safe_text(values.get(fid, default)) if fid else safe_text(default)



    def add_character(parent, parent_prefix, parent_tag, text_value):

        if safe_text(text_value) == "":

            return None

        el = ET.SubElement(parent, q(parent_prefix, parent_tag))

        cs = ET.SubElement(el, q("gco", "CharacterString"))

        cs.text = safe_text(text_value)

        return el



    def add_decimal(parent, parent_prefix, parent_tag, text_value):

        if safe_text(text_value) == "":

            return None

        el = ET.SubElement(parent, q(parent_prefix, parent_tag))

        d = ET.SubElement(el, q("gco", "Decimal"))

        d.text = safe_text(text_value)

        return el



    def add_integer(parent, parent_prefix, parent_tag, text_value):

        if safe_text(text_value) == "":

            return None

        el = ET.SubElement(parent, q(parent_prefix, parent_tag))

        i = ET.SubElement(el, q("gco", "Integer"))

        i.text = safe_text(text_value)

        return el



    def add_boolean(parent, parent_prefix, parent_tag, text_value):

        if safe_text(text_value) == "":

            return None

        el = ET.SubElement(parent, q(parent_prefix, parent_tag))

        b = ET.SubElement(el, q("gco", "Boolean"))

        b.text = "true" if safe_text(text_value).lower() in ("true", "yes", "1") else "false"

        return el



    def add_datetime(parent, parent_prefix, parent_tag, text_value):

        if safe_text(text_value) == "":

            return None

        el = ET.SubElement(parent, q(parent_prefix, parent_tag))

        dt = ET.SubElement(el, q("gco", "DateTime"))

        v = safe_text(text_value)

        if len(v) == 10 and v.count("-") == 2:

            v = v + "T00:00:00"

        dt.text = v

        return el



    def add_code(parent, parent_prefix, parent_tag, code_prefix, code_tag, value, codelist_url):

        if safe_text(value) == "":

            return None

        el = ET.SubElement(parent, q(parent_prefix, parent_tag))

        cd = ET.SubElement(el, q(code_prefix, code_tag))

        cd.set("codeList", codelist_url)

        cd.set("codeListValue", safe_text(value))

        cd.text = safe_text(value)

        return el



    def add_identifier(parent, parent_prefix, parent_tag, code_value, desc_value=""):

        if not safe_text(code_value):

            return None

        wrap = ET.SubElement(parent, q(parent_prefix, parent_tag))

        ident = ET.SubElement(wrap, q("mcc", "MD_Identifier"))

        add_character(ident, "mcc", "code", code_value)

        if safe_text(desc_value):

            add_character(ident, "mcc", "description", desc_value)

        return wrap



    def add_ci_date(parent, parent_prefix, parent_tag, date_value, date_type_value):

        if not safe_text(date_value) and not safe_text(date_type_value):

            return None

        wrap = ET.SubElement(parent, q(parent_prefix, parent_tag))

        ci_date = ET.SubElement(wrap, q("cit", "CI_Date"))

        add_datetime(ci_date, "cit", "date", date_value or now_iso())

        add_code(

            ci_date, "cit", "dateType", "cit", "CI_DateTypeCode",

            date_type_value or "creation",

            "https://schemas.isotc211.org/resources/codelists/codelists.xml#ISO19115.cit.CI_DateTypeCode"

        )

        return wrap



    def add_online_resource(contact_el, linkage):

        if not safe_text(linkage):

            return

        online = ET.SubElement(contact_el, q("cit", "onlineResource"))

        or_el = ET.SubElement(online, q("cit", "CI_OnlineResource"))

        link = ET.SubElement(or_el, q("cit", "linkage"))

        url = ET.SubElement(link, q("gcx", "Anchor"))

        url.set(q("xlink", "href"), safe_text(linkage))

        url.text = safe_text(linkage)



    def add_address(contact_el, email):

        if not safe_text(email):

            return

        addr_wrap = ET.SubElement(contact_el, q("cit", "address"))

        addr = ET.SubElement(addr_wrap, q("cit", "CI_Address"))

        mail = ET.SubElement(addr, q("cit", "electronicMailAddress"))

        cs = ET.SubElement(mail, q("gco", "CharacterString"))

        cs.text = safe_text(email)



    def add_phone(contact_el, voice):

        if not safe_text(voice):

            return

        ph_wrap = ET.SubElement(contact_el, q("cit", "phone"))

        ph = ET.SubElement(ph_wrap, q("cit", "CI_Telephone"))

        v = ET.SubElement(ph, q("cit", "number"))

        cs = ET.SubElement(v, q("gco", "CharacterString"))

        cs.text = safe_text(voice)



    def add_responsibility(parent, parent_prefix, parent_tag, role, org_name, individual_name="", position="", email="", voice="", online=""):

        if not safe_text(role) and not safe_text(org_name) and not safe_text(individual_name):

            return None

        wrap = ET.SubElement(parent, q(parent_prefix, parent_tag))

        resp = ET.SubElement(wrap, q("cit", "CI_Responsibility"))

        add_code(

            resp, "cit", "role", "cit", "CI_RoleCode", role or "pointOfContact",

            "https://schemas.isotc211.org/resources/codelists/codelists.xml#ISO19115.cit.CI_RoleCode"

        )

        party = ET.SubElement(resp, q("cit", "party"))

        if safe_text(individual_name):

            indiv = ET.SubElement(party, q("cit", "CI_Individual"))

            add_character(indiv, "cit", "name", individual_name)

            if safe_text(position):

                add_character(indiv, "cit", "positionName", position)

            if safe_text(org_name) or safe_text(email) or safe_text(voice) or safe_text(online):

                ciw = ET.SubElement(indiv, q("cit", "contactInfo"))

                contact_el = ET.SubElement(ciw, q("cit", "CI_Contact"))

                add_phone(contact_el, voice)

                add_address(contact_el, email)

                add_online_resource(contact_el, online)

        else:

            org = ET.SubElement(party, q("cit", "CI_Organisation"))

            add_character(org, "cit", "name", org_name)

            if safe_text(email) or safe_text(voice) or safe_text(online):

                ciw = ET.SubElement(org, q("cit", "contactInfo"))

                contact_el = ET.SubElement(ciw, q("cit", "CI_Contact"))

                add_phone(contact_el, voice)

                add_address(contact_el, email)

                add_online_resource(contact_el, online)

        return wrap



    root = ET.Element(q("mdb", "MD_Metadata"))

    root.set(q("xsi", "schemaLocation"),

        "https://schemas.isotc211.org/19103/-/gco/1.2 https://schemas.isotc211.org/19103/-/gco/1.2.0/gco.xsd "

        "https://schemas.isotc211.org/19103/-/gcx/1.2 https://schemas.isotc211.org/19103/-/gcx/1.2.0/gcx.xsd "

        "https://schemas.isotc211.org/19115/-1/cit/1.3 https://schemas.isotc211.org/19115/-1/cit/1.3.0/cit.xsd "

        "https://schemas.isotc211.org/19115/-1/gex/1.3 https://schemas.isotc211.org/19115/-1/gex/1.3.0/gex.xsd "

        "https://schemas.isotc211.org/19115/-1/lan/1.3 https://schemas.isotc211.org/19115/-1/lan/1.3.0/lan.xsd "

        "https://schemas.isotc211.org/19115/-1/mcc/1.3 https://schemas.isotc211.org/19115/-1/mcc/1.3.0/mcc.xsd "

        "https://schemas.isotc211.org/19115/-1/mdb/1.3 https://schemas.isotc211.org/19115/-1/mdb/1.3.0/mdb.xsd "

        "https://schemas.isotc211.org/19115/-1/mri/1.3 https://schemas.isotc211.org/19115/-1/mri/1.3.0/mri.xsd"

    )



    # Keep your interoperability register in the XML without breaking ISO parsers.

    # It is stored as XML comments near the top so ISO elements remain clean.

    interop = values.get("_interop", {}) or {}

    if interop:

        root.append(ET.Comment("SOI Interoperability Register"))

        for key in [

            "DatasetStoragePath", "MetadataXMLPath", "QGISLayerName", "ArcGISProItemName",

            "GeoServerWorkspace", "GeoServerStore", "GeoServerLayer", "ArcGISServerService",

            "ArcGISOnlineItemID"

        ]:

            if safe_text(interop.get(key, "")):

                root.append(ET.Comment(f"{key}: {safe_text(interop.get(key, ''))}"))



    add_identifier(root, "mdb", "metadataIdentifier", get("metadataIdentifier · code"), get("metadataIdentifier · description"))



    add_responsibility(

        root, "mdb", "contact",

        get("contact · role") or "custodian",

        get("contact · party name"),

        email=get("contact · electronicMailAddress"),

        online=get("contact · onlineResource linkage")

    )



    add_ci_date(root, "mdb", "dateInfo", get("dateInfo · date") or now_iso(), get("dateInfo · dateType") or "creation")



    if get("metadataScope · resourceScope"):

        ms = ET.SubElement(root, q("mdb", "metadataScope"))

        scope = ET.SubElement(ms, q("mdb", "MD_MetadataScope"))

        add_code(scope, "mdb", "resourceScope", "mcc", "MD_ScopeCode", get("metadataScope · resourceScope"),

                 "https://schemas.isotc211.org/resources/codelists/codelists.xml#ISO19115.mcc.MD_ScopeCode")



    if get("metadataStandard · title") or get("metadataStandard · edition"):

        ms = ET.SubElement(root, q("mdb", "metadataStandard"))

        citn = ET.SubElement(ms, q("cit", "CI_Citation"))

        add_character(citn, "cit", "title", get("metadataStandard · title") or "ISO 19115-1 Geographic information — Metadata")

        add_character(citn, "cit", "edition", get("metadataStandard · edition"))

        if get("metadataStandard · editionDate"):

            add_ci_date(citn, "cit", "date", get("metadataStandard · editionDate"), "edition")



    # Spatial representation - grid and vector basics.

    if any(get(x) for x in ["numberOfDimensions", "axisDimensionProperties · column · dimensionSize", "axisDimensionProperties · row · dimensionSize", "cellGeometry"]):

        sri = ET.SubElement(root, q("mdb", "spatialRepresentationInfo"))

        grid = ET.SubElement(sri, q("msr", "MD_GridSpatialRepresentation"))

        add_integer(grid, "msr", "numberOfDimensions", get("numberOfDimensions"))

        if get("axisDimensionProperties · column · dimensionSize"):

            adp = ET.SubElement(grid, q("msr", "axisDimensionProperties"))

            dim = ET.SubElement(adp, q("msr", "MD_Dimension"))

            add_code(dim, "msr", "dimensionName", "msr", "MD_DimensionNameTypeCode", "column",

                     "https://schemas.isotc211.org/resources/codelists/codelists.xml#ISO19115.msr.MD_DimensionNameTypeCode")

            add_integer(dim, "msr", "dimensionSize", get("axisDimensionProperties · column · dimensionSize"))

        if get("axisDimensionProperties · row · dimensionSize"):

            adp = ET.SubElement(grid, q("msr", "axisDimensionProperties"))

            dim = ET.SubElement(adp, q("msr", "MD_Dimension"))

            add_code(dim, "msr", "dimensionName", "msr", "MD_DimensionNameTypeCode", "row",

                     "https://schemas.isotc211.org/resources/codelists/codelists.xml#ISO19115.msr.MD_DimensionNameTypeCode")

            add_integer(dim, "msr", "dimensionSize", get("axisDimensionProperties · row · dimensionSize"))

        add_code(grid, "msr", "cellGeometry", "msr", "MD_CellGeometryCode", get("cellGeometry"),

                 "https://schemas.isotc211.org/resources/codelists/codelists.xml#ISO19115.msr.MD_CellGeometryCode")



    # CRS / reference system.

    if get("Horizontal CRS · code"):

        rsi = ET.SubElement(root, q("mdb", "referenceSystemInfo"))

        rs = ET.SubElement(rsi, q("mrs", "MD_ReferenceSystem"))

        add_identifier(rs, "mrs", "referenceSystemIdentifier", get("Horizontal CRS · code"), get("CRS description / notes"))

    if get("Vertical CRS · code"):

        rsi = ET.SubElement(root, q("mdb", "referenceSystemInfo"))

        rs = ET.SubElement(rsi, q("mrs", "MD_ReferenceSystem"))

        add_identifier(rs, "mrs", "referenceSystemIdentifier", get("Vertical CRS · code"), get("Vertical datum description"))



    # Identification info.

    ii = ET.SubElement(root, q("mdb", "identificationInfo"))

    mdi = ET.SubElement(ii, q("mri", "MD_DataIdentification"))

    citation = ET.SubElement(mdi, q("mri", "citation"))

    ci = ET.SubElement(citation, q("cit", "CI_Citation"))

    add_character(ci, "cit", "title", get("citation · title") or get("File name") or "Untitled Dataset")

    add_character(ci, "cit", "alternateTitle", get("citation · alternateTitle"))

    add_ci_date(ci, "cit", "date", get("citation · date") or now_iso(), get("citation · dateType") or "publication")

    add_identifier(ci, "cit", "identifier", get("citation · identifier · code"))

    add_character(ci, "cit", "edition", get("citation · edition"))

    if get("citation · editionDate"):

        add_ci_date(ci, "cit", "date", get("citation · editionDate"), "edition")

    if get("citation · presentationForm"):

        add_code(ci, "cit", "presentationForm", "cit", "CI_PresentationFormCode", get("citation · presentationForm"),

                 "https://schemas.isotc211.org/resources/codelists/codelists.xml#ISO19115.cit.CI_PresentationFormCode")

    if get("citation · series · name"):

        sw = ET.SubElement(ci, q("cit", "series"))

        series = ET.SubElement(sw, q("cit", "CI_Series"))

        add_character(series, "cit", "name", get("citation · series · name"))



    add_character(mdi, "mri", "abstract", get("abstract") or get("Brief description") or "No abstract provided.")

    add_character(mdi, "mri", "purpose", get("purpose"))

    add_character(mdi, "mri", "credit", get("credit"))

    add_code(mdi, "mri", "spatialRepresentationType", "mcc", "MD_SpatialRepresentationTypeCode", get("spatialRepresentationType"),

             "https://schemas.isotc211.org/resources/codelists/codelists.xml#ISO19115.mcc.MD_SpatialRepresentationTypeCode")



    if get("spatialResolution · equivalentScale"):

        sr = ET.SubElement(mdi, q("mri", "spatialResolution"))

        res = ET.SubElement(sr, q("mri", "MD_Resolution"))

        es = ET.SubElement(res, q("mri", "equivalentScale"))

        frac = ET.SubElement(es, q("mri", "MD_RepresentativeFraction"))

        add_integer(frac, "mri", "denominator", get("spatialResolution · equivalentScale"))

    if get("spatialResolution · distance (GSD)"):

        sr = ET.SubElement(mdi, q("mri", "spatialResolution"))

        res = ET.SubElement(sr, q("mri", "MD_Resolution"))

        dist = ET.SubElement(res, q("mri", "distance"))

        measure = ET.SubElement(dist, q("gco", "Distance"))

        measure.set("uom", "m")

        measure.text = safe_text(get("spatialResolution · distance (GSD)"))



    add_character(mdi, "mri", "temporalResolution", get("temporalResolution"))

    add_character(mdi, "mri", "supplementalInformation", "\n".join([x for x in [get("levelOfDetail"), get("Point cloud density / classification"), get("DEM / Geoid notes"), get("supplementalInformation")] if x]))



    if get("topicCategory"):

        tc = ET.SubElement(mdi, q("mri", "topicCategory"))

        tcc = ET.SubElement(tc, q("mri", "MD_TopicCategoryCode"))

        tcc.text = get("topicCategory")



    # Point of contact.

    add_responsibility(

        mdi, "mri", "pointOfContact",

        get("pointOfContact · role") or "pointOfContact",

        get("pointOfContact · party · organisation name"),

        get("pointOfContact · party · individual name"),

        get("pointOfContact · party · positionName"),

        get("pointOfContact · electronicMailAddress"),

        get("pointOfContact · voice"),

        get("pointOfContact · onlineResource linkage")

    )



    # Keywords.

    keyword_pairs = [

        ("keyword · theme", "theme"), ("keyword · product", "theme"), ("keyword · platform", "theme"),

        ("keyword · place", "place"), ("keyword · discipline", "discipline"),

        ("keyword · stratum", "stratum"), ("keyword · temporal", "temporal"),

    ]

    for label, ktype in keyword_pairs:

        if get(label):

            dkw = ET.SubElement(mdi, q("mri", "descriptiveKeywords"))

            kws = ET.SubElement(dkw, q("mri", "MD_Keywords"))

            for kw in [x.strip() for x in get(label).replace(";", ",").split(",") if x.strip()]:

                add_character(kws, "mri", "keyword", kw)

            add_code(kws, "mri", "type", "mri", "MD_KeywordTypeCode", ktype,

                     "https://schemas.isotc211.org/resources/codelists/codelists.xml#ISO19115.mri.MD_KeywordTypeCode")

            if get("thesaurusName · title"):

                th = ET.SubElement(kws, q("mri", "thesaurusName"))

                thci = ET.SubElement(th, q("cit", "CI_Citation"))

                add_character(thci, "cit", "title", get("thesaurusName · title"))

                if get("thesaurusName · date"):

                    add_ci_date(thci, "cit", "date", get("thesaurusName · date"), "publication")



    # Resource constraints.

    if get("accessConstraints") or get("useConstraints") or get("otherConstraints"):

        rc = ET.SubElement(mdi, q("mri", "resourceConstraints"))

        lc = ET.SubElement(rc, q("mco", "MD_LegalConstraints"))

        add_code(lc, "mco", "accessConstraints", "mco", "MD_RestrictionCode", get("accessConstraints"),

                 "https://schemas.isotc211.org/resources/codelists/codelists.xml#ISO19115.mco.MD_RestrictionCode")

        add_code(lc, "mco", "useConstraints", "mco", "MD_RestrictionCode", get("useConstraints"),

                 "https://schemas.isotc211.org/resources/codelists/codelists.xml#ISO19115.mco.MD_RestrictionCode")

        add_character(lc, "mco", "otherConstraints", get("otherConstraints"))

    if get("classification") or get("MD_SecurityConstraints · userNote"):

        rc = ET.SubElement(mdi, q("mri", "resourceConstraints"))

        sc = ET.SubElement(rc, q("mco", "MD_SecurityConstraints"))

        add_code(sc, "mco", "classification", "mco", "MD_ClassificationCode", get("classification") or "unclassified",

                 "https://schemas.isotc211.org/resources/codelists/codelists.xml#ISO19115.mco.MD_ClassificationCode")

        add_character(sc, "mco", "userNote", get("MD_SecurityConstraints · userNote"))

    if get("useLimitation"):

        rc = ET.SubElement(mdi, q("mri", "resourceConstraints"))

        cc = ET.SubElement(rc, q("mco", "MD_Constraints"))

        add_character(cc, "mco", "useLimitation", get("useLimitation"))



    # Extent.

    if any(get(x) for x in ["extent · description", "geographicElement · westBoundLongitude", "geographicElement · eastBoundLongitude", "geographicElement · southBoundLatitude", "geographicElement · northBoundLatitude"]):

        ext = ET.SubElement(mdi, q("mri", "extent"))

        ex = ET.SubElement(ext, q("gex", "EX_Extent"))

        add_character(ex, "gex", "description", get("extent · description"))

        if all(get(x) for x in ["geographicElement · westBoundLongitude", "geographicElement · eastBoundLongitude", "geographicElement · southBoundLatitude", "geographicElement · northBoundLatitude"]):

            ge = ET.SubElement(ex, q("gex", "geographicElement"))

            bbox = ET.SubElement(ge, q("gex", "EX_GeographicBoundingBox"))

            add_decimal(bbox, "gex", "westBoundLongitude", get("geographicElement · westBoundLongitude"))

            add_decimal(bbox, "gex", "eastBoundLongitude", get("geographicElement · eastBoundLongitude"))

            add_decimal(bbox, "gex", "southBoundLatitude", get("geographicElement · southBoundLatitude"))

            add_decimal(bbox, "gex", "northBoundLatitude", get("geographicElement · northBoundLatitude"))

        if get("verticalElement · minimumValue") or get("verticalElement · maximumValue"):

            ve = ET.SubElement(ex, q("gex", "verticalElement"))

            vex = ET.SubElement(ve, q("gex", "EX_VerticalExtent"))

            add_decimal(vex, "gex", "minimumValue", get("verticalElement · minimumValue"))

            add_decimal(vex, "gex", "maximumValue", get("verticalElement · maximumValue"))

        if get("temporalElement · extent"):

            te = ET.SubElement(ex, q("gex", "temporalElement"))

            tex = ET.SubElement(te, q("gex", "EX_TemporalExtent"))

            extent = ET.SubElement(tex, q("gex", "extent"))

            ti = ET.SubElement(extent, q("gml", "TimeInstant"))

            ti.set(q("gml", "id"), "temporal_extent_1")

            tp = ET.SubElement(ti, q("gml", "timePosition"))

            tp.text = get("temporalElement · extent")



    # Locale.

    loc = ET.SubElement(mdi, q("mri", "defaultLocale"))

    pt = ET.SubElement(loc, q("lan", "PT_Locale"))

    lang = ET.SubElement(pt, q("lan", "language"))

    lc = ET.SubElement(lang, q("lan", "LanguageCode"))

    lc.set("codeList", "https://www.loc.gov/standards/iso639-2/langhome.html")

    lc.set("codeListValue", "eng")

    lc.text = "English"

    enc = ET.SubElement(pt, q("lan", "characterEncoding"))

    cs = ET.SubElement(enc, q("lan", "MD_CharacterSetCode"))

    cs.set("codeList", "http://www.iana.org/assignments/character-sets")

    cs.set("codeListValue", "utf8")

    cs.text = "UTF-8"



    # Distribution.

    if any(get(x) for x in ["distributionFormat · name", "distributionFormat · version", "distributor · organisation", "transferOptions · offLine · mediumName"]):

        di = ET.SubElement(root, q("mdb", "distributionInfo"))

        dist = ET.SubElement(di, q("mrd", "MD_Distribution"))

        if get("distributionFormat · name") or get("distributionFormat · version"):

            df = ET.SubElement(dist, q("mrd", "distributionFormat"))

            fmt = ET.SubElement(df, q("mrd", "MD_Format"))

            add_character(fmt, "mrd", "name", get("distributionFormat · name"))

            add_character(fmt, "mrd", "version", get("distributionFormat · version") or "unknown")

        if get("distributor · organisation"):

            dw = ET.SubElement(dist, q("mrd", "distributor"))

            d = ET.SubElement(dw, q("mrd", "MD_Distributor"))

            add_responsibility(d, "mrd", "distributorContact", "distributor", get("distributor · organisation"), email=get("distributor · electronicMailAddress"))

        if get("transferOptions · offLine · mediumName"):

            tow = ET.SubElement(dist, q("mrd", "transferOptions"))

            to = ET.SubElement(tow, q("mrd", "MD_DigitalTransferOptions"))

            off = ET.SubElement(to, q("mrd", "offLine"))

            med = ET.SubElement(off, q("mrd", "MD_Medium"))

            add_character(med, "mrd", "name", get("transferOptions · offLine · mediumName"))



    # Lineage.

    if get("resourceLineage · statement"):

        rl = ET.SubElement(root, q("mdb", "resourceLineage"))

        lin = ET.SubElement(rl, q("mrl", "LI_Lineage"))

        add_character(lin, "mrl", "statement", get("resourceLineage · statement"))



    # Tool reload block: keeps the full UI values inside the XML so the inventory can

    # auto-fetch the linked XML and restore all fields. This is intentionally placed

    # after the ISO metadata body and can be ignored by external catalogues.

    soi_block = ET.SubElement(root, "SOI_FieldValues")

    for field in SOI_TEMPLATE_FIELDS:

        val = safe_text(values.get(field["id"], ""))

        if val:

            field_el = ET.SubElement(soi_block, "Field")

            field_el.set("id", field["id"])

            field_el.set("label", field["label"])

            field_el.text = val

    interop_block = ET.SubElement(root, "SOI_Interop")

    for key, val in (values.get("_interop", {}) or {}).items():

        if safe_text(val):

            child = ET.SubElement(interop_block, key)

            child.text = safe_text(val)


    inventory_block = ET.SubElement(root, "SOI_Inventory")
    for key, val in (values.get("_soi_inventory", {}) or {}).items():
        if safe_text(val):
            field_el = ET.SubElement(inventory_block, "Field")
            field_el.set("name", safe_text(key))
            field_el.text = safe_text(val)



    tree = ET.ElementTree(root)

    ET.indent(tree, space="  ", level=0)

    tree.write(output_xml_path, encoding="utf-8", xml_declaration=True)







# The visible SOI inventory form intentionally contains only registry fields that
# are not already represented by the ISO metadata tabs.  The five identity
# technical columns remain in the workbook for safe upserts and change checks.
SOI_INVENTORY_HEADERS = [
    "SlNO",
    "DATASET TITLE",
    "NAME OF THE DATA",
    "DATA TYPE",
    "DATASET RESPONSIBLE PARTY",
    "ORIGINAL SOURCE",
    "SURVEY SCALE",
    "SURVEY YEAR",
    "PUBLICATION YEAR",
    "AVAILABILITY",
    "LANGUAGE",
    "RIGHTS",
    "DISTRIBUTION FORMAT",
    "METADATA DATE STAMP",
    "DATA PHYSICAL LOCATION",
    "REMARKS",
]

AUTO_DERIVED_INVENTORY_HEADERS = {
    "DATASET TITLE",
    "NAME OF THE DATA",
    "DATA TYPE",
    "DATASET RESPONSIBLE PARTY",
    "DISTRIBUTION FORMAT",
    "RIGHTS",
    "METADATA DATE STAMP",
}

SYSTEM_INVENTORY_HEADERS = [
    "Dataset_Key",
    "Dataset_Hash",
    "Dataset_ID",
    "Data_Path",
    "Metadata_XML_Path",
    "Layer_Name",
    "Dataset_Modified_Time",
    "Metadata_XML_Modified_Time",
    "Last_Modified",
]

INVENTORY_HEADERS = SOI_INVENTORY_HEADERS + SYSTEM_INVENTORY_HEADERS
SOI_INVENTORY_FIELD_COUNT = len(SOI_INVENTORY_HEADERS)


def split_qgis_source(source):

    """Return the base source and QGIS provider parameters."""

    source = safe_text(source).strip()

    if "|" not in source:

        return source, []

    base, raw_params = source.split("|", 1)

    params = []

    for item in raw_params.split("|"):

        key, sep, value = item.partition("=")

        params.append((key.strip().lower(), value.strip() if sep else ""))

    return base, params


def normalized_source_path(source):

    """Normalize only the physical/service source, excluding volatile layer selectors."""

    base, _params = split_qgis_source(source)

    base = safe_text(base).strip().replace("\\", "/")

    if not base:

        return ""

    # Filesystem paths are made absolute; provider/service strings remain stable text.

    looks_like_path = bool(re.match(r"^[A-Za-z]:/|^/|^\.\.?/", base)) or os.path.exists(base)

    if looks_like_path:

        try:

            base = os.path.abspath(base).replace("\\", "/")

        except Exception:

            pass

    base = re.sub(r"/+", "/", base.rstrip("/"))

    return base.casefold()


def source_layer_name(source, fallback=""):

    """Resolve a stable layer name, preferring layername over volatile layerid."""

    _base, params = split_qgis_source(source)

    for key, value in params:

        if key == "layername" and value:

            return value.strip().casefold()

    return safe_text(fallback).strip().casefold()


def canonical_dataset_key(source, layer_name=""):

    """Return ``normalized source::layer`` independent of case and layerid."""

    source_key = normalized_source_path(source)

    layer_key = source_layer_name(source, layer_name)

    if not source_key:

        return layer_key

    return source_key + ("::" + layer_key if layer_key else "")


def dataset_hash(source, layer_name=""):

    key = canonical_dataset_key(source, layer_name)

    return hashlib.sha1(key.encode("utf-8")).hexdigest() if key else ""


def normalize_path(path):

    """Compatibility path normalizer used for legacy source/XML fallback matching."""

    return canonical_dataset_key(path)


def file_modified_iso(path):

    try:

        if path and os.path.exists(path):

            return datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%dT%H:%M:%S")

    except Exception:

        pass

    return ""


def file_size_mb(path):

    try:

        if path and os.path.exists(path):

            return str(round(os.path.getsize(path) / (1024 * 1024), 3))

    except Exception:

        pass

    return ""



def inventory_worksheet(workbook):
    """Return the named inventory worksheet, accepting legacy first-sheet files."""
    if "Metadata Inventory" in workbook.sheetnames:
        return workbook["Metadata Inventory"]
    worksheet = workbook.active
    worksheet.title = "Metadata Inventory"
    return worksheet


def read_inventory_xlsx(xlsx_path):
    """Read the Metadata Inventory worksheet using openpyxl only."""
    if not os.path.exists(xlsx_path):
        return []
    from openpyxl import load_workbook
    try:
        workbook = load_workbook(xlsx_path, data_only=True)
        worksheet = inventory_worksheet(workbook)
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            workbook.close()
            return []
        headers = [safe_text(value) for value in header_row]
        records = []
        for values in rows:
            if not values or not any(safe_text(value) for value in values):
                continue
            records.append({
                header: safe_text(values[index]) if index < len(values) else ""
                for index, header in enumerate(headers) if header
            })
        workbook.close()
        return records
    except Exception as exc:
        raise Exception(f"Could not read inventory Excel: {exc}")


def style_inventory_worksheet(worksheet, headers):
    """Apply lightweight formatting without rebuilding existing workbook content."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, header in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 2, 14), 55)


def write_inventory_xlsx(xlsx_path, records=None):
    """Create an inventory workbook; existing workbooks are never rebuilt here."""
    from openpyxl import Workbook
    xlsx_path = safe_text(xlsx_path)
    if not xlsx_path:
        raise Exception("Inventory Excel path is blank.")
    if not xlsx_path.lower().endswith(".xlsx"):
        xlsx_path = os.path.splitext(xlsx_path)[0] + ".xlsx"
    folder = os.path.dirname(xlsx_path)
    if folder:
        os.makedirs(folder, exist_ok=True)
    if os.path.exists(xlsx_path):
        return xlsx_path
    records = records or []
    headers = INVENTORY_HEADERS[:]
    for record in records:
        for key in record:
            if key not in headers:
                headers.append(key)
    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Metadata Inventory"
        worksheet.append(headers)
        for record in records:
            worksheet.append([safe_text(record.get(header, "")) for header in headers])
        style_inventory_worksheet(worksheet, headers)
        workbook.save(xlsx_path)
        workbook.close()
        return xlsx_path
    except PermissionError:
        raise Exception(f"Inventory Excel is open or locked. Close the workbook and try again.\n\nLocked file: {xlsx_path}")
    except Exception as exc:
        raise Exception(f"Could not create inventory Excel: {exc}")


def inventory_identity(record):
    """Return normalized values for every supported inventory identity column."""
    data_path = safe_text(record.get("Data_Path") or record.get("Dataset_Path") or record.get("Dataset_Source"))
    layer_name = safe_text(record.get("Layer_Name"))
    return {
        "Dataset_Key": safe_text(record.get("Dataset_Key")) or canonical_dataset_key(data_path, layer_name),
        "Dataset_Hash": safe_text(record.get("Dataset_Hash")),
        "Dataset_ID": safe_text(record.get("Dataset_ID") or record.get("Dataset_GUID")),
        "Data_Path": data_path,
        "Metadata_XML_Path": safe_text(record.get("Metadata_XML_Path") or record.get("XML_Path")),
        "Layer_Name": layer_name,
    }


def find_inventory_record(records, data_path="", xml_path="", dataset_id="", dataset_key="", dataset_hash_value="", layer_name=""):
    """Find a record using each supported stable identifier, strongest first."""
    requested = {
        "Dataset_Key": safe_text(dataset_key) or canonical_dataset_key(data_path, layer_name),
        "Dataset_Hash": safe_text(dataset_hash_value) or dataset_hash(data_path, layer_name),
        "Dataset_ID": safe_text(dataset_id),
        "Data_Path": canonical_dataset_key(data_path, layer_name),
        "Metadata_XML_Path": normalized_source_path(xml_path),
    }

    def matches(record, field):
        identity = inventory_identity(record)
        wanted = requested[field]
        if not wanted:
            return False
        if field == "Data_Path":
            actual = canonical_dataset_key(identity[field], identity["Layer_Name"])
        elif field == "Metadata_XML_Path":
            actual = normalized_source_path(identity[field])
        else:
            actual = identity[field]
        return safe_text(actual).casefold() == safe_text(wanted).casefold()

    for field in ("Dataset_Key", "Dataset_Hash", "Dataset_ID", "Data_Path", "Metadata_XML_Path"):
        for index, record in enumerate(records):
            if matches(record, field):
                return index, record
    return None, None


def upsert_inventory_xlsx(xlsx_path, row):
    """Create or update one row in place using openpyxl only."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment

    xlsx_path = safe_text(xlsx_path)
    if not xlsx_path:
        raise Exception("Inventory Excel path is blank.")
    if not xlsx_path.lower().endswith(".xlsx"):
        xlsx_path = os.path.splitext(xlsx_path)[0] + ".xlsx"
    if not os.path.exists(xlsx_path):
        write_inventory_xlsx(xlsx_path)

    row = {safe_text(key): safe_text(value) for key, value in row.items() if safe_text(key)}
    source = row.get("Data_Path", "")
    layer_name = row.get("Layer_Name", "")
    row["Dataset_Key"] = row.get("Dataset_Key") or canonical_dataset_key(source, layer_name)
    row["Dataset_Hash"] = row.get("Dataset_Hash") or dataset_hash(source, layer_name)
    row["Dataset_ID"] = row.get("Dataset_ID", "")
    row["Metadata_XML_Path"] = row.get("Metadata_XML_Path", "")
    row["Last_Modified"] = row.get("Last_Modified") or now_iso()

    workbook = None
    try:
        workbook = load_workbook(xlsx_path)
        worksheet = inventory_worksheet(workbook)
        headers = [safe_text(cell.value) for cell in worksheet[1]] if worksheet.max_row else []
        if not any(headers):
            headers = INVENTORY_HEADERS[:]
            for column, header in enumerate(headers, start=1):
                worksheet.cell(1, column, header)
        for header in INVENTORY_HEADERS:
            if header not in headers:
                headers.append(header)
                worksheet.cell(1, len(headers), header)

        records = []
        excel_rows = []
        for excel_row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if values and any(safe_text(value) for value in values):
                records.append({header: safe_text(values[i]) if i < len(values) else "" for i, header in enumerate(headers) if header})
                excel_rows.append(excel_row_number)

        index, existing = find_inventory_record(
            records,
            data_path=source,
            xml_path=row["Metadata_XML_Path"],
            dataset_id=row["Dataset_ID"],
            dataset_key=row["Dataset_Key"],
            dataset_hash_value=row["Dataset_Hash"],
            layer_name=layer_name,
        )
        if index is None:
            excel_row = worksheet.max_row + 1
            action = "created" if not records else "added"
            row["SlNO"] = row.get("SlNO") or str(len(records) + 1)
        else:
            excel_row = excel_rows[index]
            action = "updated"
            row["SlNO"] = row.get("SlNO") or safe_text((existing or {}).get("SlNO")) or str(index + 1)

        columns = {header: position for position, header in enumerate(headers, start=1)}
        for header in INVENTORY_HEADERS:
            cell = worksheet.cell(excel_row, columns[header])
            cell.value = row.get(header, "")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        style_inventory_worksheet(worksheet, headers)
        workbook.save(xlsx_path)
        workbook.close()
        workbook = None
    except PermissionError:
        raise Exception(
            "Inventory Excel is open or locked in Excel. Close the workbook and try again."
            f"\n\nLocked file: {xlsx_path}"
        )
    except Exception as exc:
        if "permission" in safe_text(exc).lower() or "denied" in safe_text(exc).lower():
            raise Exception(
                "Inventory Excel is open or locked in Excel. Close the workbook and try again."
                f"\n\nLocked file: {xlsx_path}"
            )
        raise Exception(f"Could not update inventory Excel: {exc}")
    finally:
        if workbook is not None:
            workbook.close()

    total_records = len(read_inventory_xlsx(xlsx_path))
    return action, xlsx_path, excel_row, total_records


def read_metadata_xml_values(xml_path):

    values = {}

    interop = {}
    inventory = {}

    if not xml_path or not os.path.exists(xml_path):
        return values, interop, inventory

    root = ET.parse(xml_path).getroot()

    label_to_id = {f["label"]: f["id"] for f in SOI_TEMPLATE_FIELDS}



    # Older custom XML format used label attributes directly.

    for el in root.iter():

        label = el.attrib.get("label")

        fid = el.attrib.get("id")

        if fid and fid.startswith("f_"):

            values[fid] = safe_text(el.text)

        elif label and label in label_to_id:

            values[label_to_id[label]] = safe_text(el.text)



    # New ISO XML writer stores a reload block so the GUI can be restored from inventory-linked XML.

    for field_el in root.findall(".//Field"):

        fid = field_el.attrib.get("id", "")

        label = field_el.attrib.get("label", "")

        if fid:

            values[fid] = safe_text(field_el.text)

        elif label in label_to_id:

            values[label_to_id[label]] = safe_text(field_el.text)



    for tag_name in ("Interoperability", "SOI_Interop"):

        interop_el = root.find(tag_name)

        if interop_el is not None:

            for child in list(interop_el):

                interop[child.tag] = safe_text(child.text)



    inventory_el = root.find("SOI_Inventory")
    if inventory_el is not None:
        for field_el in inventory_el.findall("Field"):
            name = safe_text(field_el.attrib.get("name"))
            if name:
                inventory[name] = safe_text(field_el.text)

    return values, interop, inventory





def compare_records(old_values, new_values, label_map=None, ignore_blank_old=False):

    changes = []

    all_keys = sorted(set(old_values.keys()) | set(new_values.keys()))

    for key in all_keys:

        old = safe_text(old_values.get(key, ""))

        new = safe_text(new_values.get(key, ""))

        if ignore_blank_old and old == "":

            continue

        if old != new:

            label = label_map.get(key, key) if label_map else key

            changes.append((label, old, new))

    return changes





class SOIMetadataDialog(QDialog):

    def __init__(self):

        super().__init__()

        self.setWindowTitle("SOI ISO 19115-1 Metadata XML Generator - QGIS")

        self.resize(1200, 860)

        # Global UI style.

        # Important: QComboBox QLineEdit is styled explicitly because editable combo boxes

        # otherwise show a white strip inside the selected-value area on Windows/QGIS Qt themes.

        self.setStyleSheet("""

            QDialog { background-color: #f7f9fb; }

            QLabel { font-size: 12px; color: #263238; }

            QGroupBox {

                font-weight: bold;

                border: 1px solid #cfd8dc;

                border-radius: 6px;

                margin-top: 8px;

                padding: 8px;

                background-color: #ffffff;

            }

            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 4px; }

            QLineEdit, QTextEdit, QComboBox {

                border: 1px solid #b0bec5;

                border-radius: 4px;

                padding: 5px;

                background-color: #ffffff;

                color: #111111;

                selection-background-color: #bbdefb;

                selection-color: #000000;

            }

            QComboBox {

                min-height: 22px;

                padding-right: 24px;

            }

            QComboBox QLineEdit {

                background-color: #ffffff;

                color: #111111;

                border: 0px;

                padding: 0px 2px;

                margin: 0px;

            }

            QComboBox::drop-down {

                subcontrol-origin: padding;

                subcontrol-position: top right;

                width: 22px;

                border-left: 1px solid #b0bec5;

                border-top-right-radius: 4px;

                border-bottom-right-radius: 4px;

                background-color: #eef5f9;

            }

            QComboBox::down-arrow {

                width: 0px;

                height: 0px;

            }

            QComboBox QAbstractItemView {

                background-color: #ffffff;

                color: #111111;

                border: 1px solid #90a4ae;

                selection-background-color: #e3f2fd;

                selection-color: #000000;

                outline: 0px;

            }

            QPushButton {

                padding: 7px 12px;

                font-weight: bold;

                border-radius: 5px;

                background-color: #e3f2fd;

                border: 1px solid #90caf9;

                color: #0d47a1;

            }

            QPushButton:hover { background-color: #bbdefb; }

            QPushButton:disabled { background-color: #eceff1; color: #78909c; border: 1px solid #cfd8dc; }

            QTabWidget::pane { border: 1px solid #cfd8dc; background-color: #ffffff; }

            QTabBar::tab { padding: 8px 12px; margin-right: 2px; background-color: #f1f4f6; border: 1px solid #cfd8dc; border-bottom: none; }

            QTabBar::tab:selected { background-color: #e3f2fd; font-weight: bold; }

            QScrollArea { background-color: #ffffff; border: none; }

        """)



        self.widgets = {}

        self.field_lookup = {f["label"]: f["id"] for f in SOI_TEMPLATE_FIELDS}

        self.current_raster_metadata = {}



        main_layout = QVBoxLayout(self)



        header = QLabel("SOI ISO 19115-1 Metadata XML Generator — QGIS (SOI Inventory Aligned v13)")

        header.setAlignment(Qt.AlignCenter)

        header.setStyleSheet("font-size:18px; font-weight:bold; padding:6px;")

        main_layout.addWidget(header)



        info = QLabel(

            "Workflow: 1) Select a layer or dataset  2) Extract auto metadata  3) Fill mandatory ISO fields  4) Generate ISO XML + update inventory. SOI Inventory Record and Interoperability Register are available as scrollable tabs; bottom action buttons remain fixed."

        )

        info.setStyleSheet("color:#555; padding-bottom:4px;")

        main_layout.addWidget(info)



        file_box = QGroupBox("Dataset Selection")

        file_layout = QVBoxLayout(file_box)



        source_row = QHBoxLayout()

        source_row.addWidget(QLabel("Data Source Type"))

        self.source_type_combo = QComboBox()

        self.source_type_combo.addItems([

            "Layer from Current QGIS Project (Recommended)",

            "File System Dataset",

            "File Geodatabase (.gdb) — Vector/Raster",

            "GeoPackage (.gpkg) — Vector/Raster",

        ])

        self.source_type_combo.currentIndexChanged.connect(self.set_source_mode)

        source_row.addWidget(self.source_type_combo)

        source_row.addStretch()

        file_layout.addLayout(source_row)



        self.project_layer_row = QHBoxLayout()

        self.project_layer_combo = QComboBox()

        self.project_layer_combo.currentIndexChanged.connect(self.on_dataset_selection_changed)

        refresh_btn = QPushButton("Refresh Project Layers")

        refresh_btn.clicked.connect(self.refresh_project_layers)

        self.project_layer_row.addWidget(QLabel("Project Layer"))

        self.project_layer_row.addWidget(self.project_layer_combo)

        self.project_layer_row.addWidget(refresh_btn)

        file_layout.addLayout(self.project_layer_row)



        self.file_path_row = QHBoxLayout()

        self.raster_path_edit = QLineEdit()

        self.raster_path_edit.setPlaceholderText("Select dataset: .shp, .tif, .tiff, .img, .jp2, .vrt, .geojson, .gpkg")

        self.raster_path_edit.editingFinished.connect(self.on_dataset_selection_changed)

        browse_btn = QPushButton("Browse File")

        browse_btn.clicked.connect(self.browse_raster)

        self.file_path_row.addWidget(QLabel("Dataset File"))

        self.file_path_row.addWidget(self.raster_path_edit)

        self.file_path_row.addWidget(browse_btn)

        file_layout.addLayout(self.file_path_row)



        self.container_row = QHBoxLayout()

        self.container_path_edit = QLineEdit()

        self.container_path_edit.setPlaceholderText("Select .gdb folder or .gpkg file")

        self.container_path_edit.editingFinished.connect(self.on_dataset_selection_changed)

        browse_container_btn = QPushButton("Browse Container")

        browse_container_btn.clicked.connect(self.browse_container)

        self.container_row.addWidget(QLabel("Container"))

        self.container_row.addWidget(self.container_path_edit)

        self.container_row.addWidget(browse_container_btn)

        file_layout.addLayout(self.container_row)



        self.container_layer_row = QHBoxLayout()

        self.container_layer_combo = QComboBox()

        self.container_layer_combo.currentIndexChanged.connect(self.on_dataset_selection_changed)

        load_container_btn = QPushButton("Load Selected Layer into QGIS")

        load_container_btn.clicked.connect(self.load_selected_container_layer)

        self.container_layer_row.addWidget(QLabel("Layer/Raster inside container"))

        self.container_layer_row.addWidget(self.container_layer_combo)

        self.container_layer_row.addWidget(load_container_btn)

        file_layout.addLayout(self.container_layer_row)



        action_row = QHBoxLayout()

        extract_btn = QPushButton("Extract Auto Metadata")

        extract_btn.clicked.connect(self.extract_and_fill)



        self.unlock_checkbox = QCheckBox("Unlock auto-derived fields")

        self.unlock_checkbox.stateChanged.connect(self.apply_readonly_rules)



        action_row.addWidget(extract_btn)

        action_row.addWidget(self.unlock_checkbox)

        action_row.addStretch()

        file_layout.addLayout(action_row)



        main_layout.addWidget(file_box)



        self.tabs = QTabWidget()

        main_layout.addWidget(self.tabs)



        self.create_interoperability_tab()

        self.create_metadata_tabs()



        bottom_row = QHBoxLayout()



        save_xml_btn = QPushButton("Generate ISO XML + Update Inventory")

        # QPushButton.clicked emits a boolean ``checked`` argument. Connecting it
        # directly to generate_xml_and_inventory(update_inventory=True) replaces
        # that default with False for a normal button click, silently selecting
        # the XML-only path. Route both buttons through explicit slots instead.
        save_xml_btn.clicked.connect(self.generate_xml_and_update_inventory)



        save_xml_only_btn = QPushButton("Generate ISO XML Only")

        save_xml_only_btn.clicked.connect(self.generate_xml_only)



        close_btn = QPushButton("Close")

        close_btn.clicked.connect(self.close)



        bottom_row.addStretch()

        bottom_row.addWidget(save_xml_only_btn)

        bottom_row.addWidget(save_xml_btn)

        bottom_row.addWidget(close_btn)



        main_layout.addLayout(bottom_row)



        # Startup must remain blank: do not auto-select or auto-load metadata.

        self._initializing = True

        self.refresh_project_layers(select_blank=True)

        self.set_source_mode()

        self.apply_readonly_rules()

        # Fixed inventory path: set once at startup and keep it unchanged when layers change.

        self.set_fixed_inventory_path_at_startup()

        # Create the register immediately if it does not exist, so the user can

        # see the actual Metadata Inventory Excel before generating XML.

        self.ensure_inventory_file_exists(show_message=False)

        self.set_status("No dataset selected; inventory register is ready", None)

        self._initializing = False



    def create_metadata_tabs(self):

        """Build task-oriented tabs instead of exposing the raw ISO section order."""

        tab_groups = [

            ("Identification", {"Dataset File Information", "MD_Metadata › Identifier", "MD_DataIdentification › General"}),

            ("Citation", {"MD_DataIdentification › Citation"}),

            ("Contact", {"MD_Metadata › Contact", "CI_Responsibility › Point of Contact"}),

            ("Keywords", {"MD_Keywords"}),

            ("Extent", {"EX_Extent"}),

            ("CRS", {"RS_ReferenceSystem › CRS", "MD_SpatialRepresentation"}),

            ("Quality", {"DQ_DataQuality"}),

            ("Constraints", {"MD_Constraints"}),

            ("Distribution", {"MD_Distribution"}),

            ("Lineage", {"LI_Lineage & Maintenance"}),

            ("Advanced", {"MD_Metadata › Date & Standard", "MD_Metadata › Scope", "MD_ContentInformation", "FC_FeatureCatalogue", "Advanced / 3D"}),

        ]

        for tab_name, sections in tab_groups:

            fields = [field for field in SOI_TEMPLATE_FIELDS if field["section"] in sections]

            if not fields:

                continue

            scroll = QScrollArea()

            scroll.setWidgetResizable(True)

            body = QWidget()

            form = QFormLayout(body)

            form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)

            current_section = None

            for field in fields:

                if field["section"] != current_section:

                    current_section = field["section"]

                    heading = QLabel(current_section)

                    heading.setStyleSheet("font-weight:bold; color:#1f4e79; padding-top:8px;")

                    form.addRow(heading)

                label_text = field["label"] + (" *" if field["required"].lower() == "yes" else "")

                label = QLabel(label_text)

                label.setToolTip(field["iso_path"])

                widget = self.create_widget_for_field(field)

                widget.setToolTip(field["iso_path"])

                self.widgets[field["id"]] = widget

                form.addRow(label, widget)

            scroll.setWidget(body)

            self.tabs.addTab(scroll, tab_name)


    def create_widget_for_field(self, field):

        label = field["label"]

        type_ = field["type"]



        if type_ == "CodeList" and label in CODELIST_VALUES:

            combo = QComboBox()

            combo.setEditable(True)  # editable so SOI can add local code values if needed

            combo.setInsertPolicy(QComboBox.NoInsert)

            combo.addItem("")

            combo.addItems(CODELIST_VALUES[label])

            # Fix editable-combo selected-value white strip in some QGIS/Qt themes.

            if combo.lineEdit():

                combo.lineEdit().setStyleSheet("background-color:#ffffff; color:#111111; border:0px; padding:0px 2px; margin:0px;")

            return combo



        if is_multiline_field(label, field["iso_path"]):

            text = QTextEdit()

            text.setMinimumHeight(70)

            return text



        return QLineEdit()



    def create_interoperability_tab(self):
        tab = QWidget()
        outer = QVBoxLayout(tab)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        fields_body = QWidget()
        form = QFormLayout(fields_body)
        self.interop_widgets = {}

        interop_fields = [
            ("DatasetStoragePath", "Dataset Storage Path"),
            ("MetadataXMLPath", "Metadata XML Path"),
            ("QGISLayerName", "QGIS Layer Name"),
            ("ArcGISProItemName", "ArcGIS Pro Item Name"),
            ("GeoServerWorkspace", "GeoServer Workspace"),
            ("GeoServerStore", "GeoServer Store"),
            ("GeoServerLayer", "GeoServer Layer"),
            ("ArcGISServerService", "ArcGIS Server Service"),
            ("ArcGISOnlineItemID", "ArcGIS Online Item ID"),
            ("InventoryExcelPath", "Metadata Inventory Excel Path (.xlsx)"),
        ]
        for key, label in interop_fields:
            edit = QLineEdit()
            self.interop_widgets[key] = edit
            form.addRow(QLabel(label), edit)

        self.registry_status_label = QLabel("Registry Status: —")
        self.registry_status_label.setWordWrap(True)
        self.registry_status_label.setStyleSheet("font-weight:bold; color:#546e7a; padding:4px;")
        form.addRow(QLabel("Auto Registry Check"), self.registry_status_label)
        scroll.setWidget(fields_body)
        outer.addWidget(scroll)

        actions = QHBoxLayout()
        for text, handler in (
            ("Set Inventory Excel", self.browse_inventory_excel),
            ("Browse XML", self.browse_existing_xml),
            ("Load XML", self.load_xml_into_gui),
            ("Update Inventory from XML", self.add_or_update_inventory_from_existing_xml),
            ("Check Changes", self.check_inventory_xml_changes),
        ):
            button = QPushButton(text)
            button.clicked.connect(handler)
            actions.addWidget(button)
        outer.addLayout(actions)

        self.tabs.addTab(tab, "Registry / Interoperability")
        self.create_soi_inventory_tab()

    def create_soi_inventory_tab(self):
        """Build the reduced SOI registry form without duplicating ISO sections."""
        container = QWidget()
        outer = QVBoxLayout(container)
        hint = QLabel(
            "Only SOI registry-specific columns are shown here. Values sourced from ISO metadata "
            "are grey and read-only unless 'Unlock auto-derived fields' is checked."
        )
        hint.setWordWrap(True)
        outer.addWidget(hint)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        body = QWidget()
        form = QFormLayout(body)
        form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        self.soi_inventory_widgets = {}
        multiline_headers = {"ORIGINAL SOURCE", "RIGHTS", "DATA PHYSICAL LOCATION", "REMARKS"}
        for header in SOI_INVENTORY_HEADERS:
            if header == "SlNO":
                note = QLabel("Assigned automatically during inventory save")
                note.setStyleSheet("color:#607d8b;")
                form.addRow(QLabel(header), note)
                continue
            widget = QTextEdit() if header in multiline_headers else QLineEdit()
            if isinstance(widget, QTextEdit):
                widget.setMinimumHeight(60)
            widget.setToolTip("Stored in Excel registry column: " + header)
            self.soi_inventory_widgets[header] = widget
            form.addRow(QLabel(header), widget)
        scroll.setWidget(body)
        outer.addWidget(scroll)
        self.tabs.addTab(container, "SOI Inventory")

    def set_soi_inventory_field(self, header, value, overwrite=False):

        if not hasattr(self, "soi_inventory_widgets"):

            return

        widget = self.soi_inventory_widgets.get(header)

        if widget is None:

            return

        current = self.get_widget_value(widget)

        if current and not overwrite:

            return

        if isinstance(widget, QTextEdit):

            widget.setPlainText(safe_text(value))

        else:

            widget.setText(safe_text(value))



    def get_soi_inventory_value(self, header):

        if not hasattr(self, "soi_inventory_widgets"):

            return ""

        widget = self.soi_inventory_widgets.get(header)

        if widget is None:

            return ""

        return self.get_widget_value(widget)



    def populate_soi_inventory_defaults(self, metadata=None, layer=None, source_path="", overwrite=False):
        """Fill reduced inventory fields from ISO values and the selected dataset."""
        metadata = metadata or getattr(self, "current_raster_metadata", {}) or {}
        source_path = source_path or metadata.get("file_path", "")
        if not source_path and hasattr(self, "interop_widgets"):
            source_path = self.interop_widgets["DatasetStoragePath"].text()

        def g(label):
            return self.get_field_by_label(label)

        layer_type = safe_text(metadata.get("layer_type"))
        responsible_party = (
            g("pointOfContact · party · organisation name")
            or g("contact · party name")
            or g("pointOfContact · party · individual name")
        )
        citation_date = g("citation · date")
        defaults = {
            "DATASET TITLE": g("citation · title"),
            "NAME OF THE DATA": g("File name") or metadata.get("File name", "") or os.path.basename(split_qgis_source(source_path)[0]),
            "DATA TYPE": "VECTOR" if layer_type.lower() == "vector" else "RASTER" if layer_type.lower() == "raster" else layer_type.upper(),
            "DATASET RESPONSIBLE PARTY": responsible_party,
            "ORIGINAL SOURCE": g("resourceLineage · statement"),
            "SURVEY SCALE": g("spatialResolution · equivalentScale"),
            "SURVEY YEAR": extract_year(citation_date),
            "PUBLICATION YEAR": extract_year(citation_date),
            "AVAILABILITY": "YES",
            "LANGUAGE": g("defaultLocale · language") or "English",
            "RIGHTS": g("useLimitation") or g("otherConstraints"),
            "DISTRIBUTION FORMAT": g("distributionFormat · name"),
            "METADATA DATE STAMP": g("dateInfo · date"),
            "DATA PHYSICAL LOCATION": source_path,
        }
        for header, value in defaults.items():
            self.set_soi_inventory_field(header, value, overwrite=overwrite or header in AUTO_DERIVED_INVENTORY_HEADERS)
        self.apply_readonly_rules()

    def clear_soi_inventory_fields(self):

        """Clear only the SOI inventory override tab values. ISO metadata fields are not touched."""

        for widget in getattr(self, "soi_inventory_widgets", {}).values():

            if isinstance(widget, QTextEdit):

                widget.clear()

            else:

                widget.setText("")

        self.set_status("SOI inventory override fields cleared", None)



    def _set_layout_visible(self, layout, visible):

        for i in range(layout.count()):

            item = layout.itemAt(i)

            w = item.widget()

            if w is not None:

                w.setVisible(visible)



    def set_source_mode(self):

        idx = self.source_type_combo.currentIndex() if hasattr(self, "source_type_combo") else 0

        self._set_layout_visible(self.project_layer_row, idx == 0)

        self._set_layout_visible(self.file_path_row, idx == 1)

        self._set_layout_visible(self.container_row, idx in (2, 3))

        self._set_layout_visible(self.container_layer_row, idx in (2, 3))



    def refresh_project_layers(self, select_blank=True):

        """Populate project-layer dropdown without selecting a real layer on startup/refresh."""

        self.project_layer_combo.blockSignals(True)

        self.project_layer_combo.clear()

        self.project_layer_combo.addItem("-- Select layer from current QGIS project --", None)



        layers = list(QgsProject.instance().mapLayers().values())

        if not layers:

            self.project_layer_combo.addItem("No layers loaded in current QGIS project", None)

        else:

            for layer in layers:

                try:

                    layer_type = "Raster" if layer.type() == QgsMapLayer.RasterLayer else "Vector" if layer.type() == QgsMapLayer.VectorLayer else "Other"

                    self.project_layer_combo.addItem(f"{layer.name()}  [{layer_type} | {layer.providerType()}]", layer)

                except Exception:

                    self.project_layer_combo.addItem(layer.name(), layer)



        self.project_layer_combo.setCurrentIndex(0 if select_blank else self.project_layer_combo.currentIndex())

        self.project_layer_combo.blockSignals(False)

        self.set_status("No dataset selected", None)



    def browse_raster(self):

        file_path, _ = QFileDialog.getOpenFileName(

            self,

            "Select Dataset",

            "",

            "GIS Datasets (*.tif *.tiff *.img *.jp2 *.vrt *.shp *.geojson *.kml *.csv *.gpkg);;All Files (*.*)"

        )



        if file_path:

            self.raster_path_edit.setText(file_path)

            self.interop_widgets["DatasetStoragePath"].setText(file_path)

            base = os.path.splitext(os.path.basename(file_path))[0]

            self.interop_widgets["QGISLayerName"].setText(base)

            self.ensure_inventory_path_exists_in_ui()

            self.on_dataset_selection_changed()



    def browse_container(self):

        mode = self.source_type_combo.currentIndex()

        if mode == 2:

            path = QFileDialog.getExistingDirectory(self, "Select File Geodatabase Folder (.gdb)", "")

        else:

            path, _ = QFileDialog.getOpenFileName(self, "Select GeoPackage", "", "GeoPackage (*.gpkg);;All Files (*.*)")



        if not path:

            return



        self.container_path_edit.setText(path)

        self.interop_widgets["DatasetStoragePath"].setText(path)

        self.ensure_inventory_path_exists_in_ui()

        self.populate_container_layers(path)

        # Do not auto-check until a real internal layer/raster is selected.

        self.set_status("Container selected. Choose a layer/raster to check registry.", None)



    def populate_container_layers(self, path):

        self.container_layer_combo.blockSignals(True)

        self.container_layer_combo.clear()

        self.container_layer_combo.addItem("-- Select layer/raster inside container --", None)

        datasets = list_container_datasets(path)



        if not datasets:

            self.container_layer_combo.addItem(

                "No layers/raster datasets found. Add it manually in QGIS and use Project Layer mode.",

                None

            )

            self.container_layer_combo.setCurrentIndex(0)

            self.container_layer_combo.blockSignals(False)

            self.set_status("Container selected, but no internal layer selected", None)

            return



        # Show vectors first and rasters second, but keep both selectable.

        for item in datasets:

            dtype = "Raster" if item.get("type") == "raster" else "Vector"

            self.container_layer_combo.addItem(f"{item.get('name')}  [{dtype}]", item)



        self.container_layer_combo.setCurrentIndex(0)

        self.container_layer_combo.blockSignals(False)

        self.set_status("Container selected. Choose a layer/raster to check registry.", None)



    def load_selected_container_layer(self):

        container_path = self.container_path_edit.text().strip()

        item = self.container_layer_combo.currentData()



        if not container_path or not item:

            QMessageBox.warning(self, "Missing Layer", "Please select a .gdb/.gpkg and a layer/raster inside it.")

            return



        layer_name = item.get("name", "Selected layer")

        uri = item.get("uri", "")

        item_type = item.get("type", "vector")



        if item_type == "raster":

            layer = QgsRasterLayer(uri, layer_name)

        else:

            layer = QgsVectorLayer(uri, layer_name, "ogr")



        if not layer.isValid():

            QMessageBox.critical(

                self,

                "Layer Load Failed",

                "QGIS could not load the selected item.\n\n"

                "For FileGDB rasters, this tool uses GDAL subdataset URIs. "

                "If it still fails, your QGIS/GDAL build may not support FileGDB raster reading."

            )

            return



        QgsProject.instance().addMapLayer(layer)

        self.refresh_project_layers()

        self.source_type_combo.setCurrentIndex(0)



        for i in range(self.project_layer_combo.count()):

            if self.project_layer_combo.itemData(i) == layer:

                self.project_layer_combo.setCurrentIndex(i)

                break



        self.interop_widgets["DatasetStoragePath"].setText(uri)

        self.interop_widgets["QGISLayerName"].setText(layer.name())

        QMessageBox.information(self, "Layer Loaded", f"Loaded into QGIS project:\n\n{layer.name()}")





    def current_selected_layer_or_path(self):

        mode = self.source_type_combo.currentIndex()



        if mode == 0:

            layer = self.project_layer_combo.currentData()

            try:

                return layer, layer.source() if layer else ""

            except Exception:

                return layer, ""



        if mode == 1:

            path = self.raster_path_edit.text().strip()

            if not path:

                return None, ""



            raster = QgsRasterLayer(path, os.path.basename(path))

            if raster.isValid():

                return raster, path



            vector = QgsVectorLayer(path, os.path.splitext(os.path.basename(path))[0], "ogr")

            if vector.isValid():

                return vector, path



            return None, path



        # Container mode: supports both vector feature classes and raster subdatasets.

        item = self.container_layer_combo.currentData()

        if item:

            uri = item.get("uri", "")

            name = item.get("name", "Selected layer")

            item_type = item.get("type", "vector")



            if item_type == "raster":

                raster = QgsRasterLayer(uri, name)

                if raster.isValid():

                    return raster, uri

            else:

                vector = QgsVectorLayer(uri, name, "ogr")

                if vector.isValid():

                    return vector, uri



        return None, self.container_path_edit.text().strip()



    def clean_source_for_filesystem(self, source_path):

        """Return a real filesystem path/folder from QGIS/GDAL source strings."""

        src = safe_text(source_path)

        if not src:

            return ""

        if "|" in src:

            src = src.split("|", 1)[0]

        # GDAL FileGDB subdataset examples can contain quoted gdb path.

        m = re.search(r'"([^"]+\.gdb)"', src, flags=re.IGNORECASE)

        if m:

            return m.group(1)

        if src.lower().startswith("openfilegdb:"):

            parts = src.split(":")

            for part in parts:

                part = part.strip('"')

                if part.lower().endswith(".gdb"):

                    return part

        return src



    def default_inventory_path_for_source(self, source_path):

        fs = self.clean_source_for_filesystem(source_path)

        if fs and os.path.exists(fs):

            folder = fs if os.path.isdir(fs) else os.path.dirname(fs)

        else:

            folder = os.path.dirname(fs) if fs else os.path.expanduser("~")

        return os.path.join(folder or os.path.expanduser("~"), "metadata_inventory_register.xlsx")



    def fixed_inventory_default_path(self):

        """Return one stable inventory register path for the tool.

        It does not depend on the selected layer/dataset, so the register path does not jump

        when the user changes layers. Preference order:

        1) Previously saved SOI tool setting

        2) Current QGIS project folder

        3) User Documents/Home folder

        """

        settings = QSettings()

        saved = safe_text(settings.value("SOI_MetadataGenerator/fixed_inventory_path", ""))

        if saved:

            return saved



        try:

            project_file = QgsProject.instance().fileName()

            if project_file:

                base_folder = os.path.dirname(project_file)

            else:

                base_folder = os.path.join(os.path.expanduser("~"), "Documents")

                if not os.path.isdir(base_folder):

                    base_folder = os.path.expanduser("~")

        except Exception:

            base_folder = os.path.expanduser("~")



        return os.path.join(base_folder, "SOI_Metadata_Register", "metadata_inventory_register.xlsx")



    def set_fixed_inventory_path_at_startup(self):

        """Set register path once. Do not call this on every layer change."""

        if not hasattr(self, "interop_widgets"):

            return

        current = safe_text(self.interop_widgets["InventoryExcelPath"].text())

        if not current:

            self.interop_widgets["InventoryExcelPath"].setText(self.fixed_inventory_default_path())



    def ensure_inventory_path_exists_in_ui(self):

        """Keep the already-selected register path. Only fill it if it is blank, then ensure the file exists."""

        if not hasattr(self, "interop_widgets"):

            return

        if not safe_text(self.interop_widgets["InventoryExcelPath"].text()):

            self.interop_widgets["InventoryExcelPath"].setText(self.fixed_inventory_default_path())

        self.ensure_inventory_file_exists(show_message=False)



    def normalize_inventory_path(self, inventory_path):

        inventory_path = safe_text(inventory_path) or self.fixed_inventory_default_path()

        if not inventory_path.lower().endswith(".xlsx"):

            inventory_path = os.path.splitext(inventory_path)[0] + ".xlsx"

        folder = os.path.dirname(inventory_path)

        if folder and not os.path.exists(folder):

            os.makedirs(folder, exist_ok=True)

        return inventory_path



    def ensure_inventory_file_exists(self, show_message=False):

        """Create the fixed Metadata Inventory Excel immediately if it is missing.



        Earlier versions only created the inventory when the user generated XML.

        This made the Interoperability Register path visible in the UI but no

        actual .xlsx file existed on disk. This method guarantees that the

        register exists as soon as the tool opens, when a path is selected, and

        before any inventory lookup/update is attempted.

        """

        try:

            inventory_path = self.normalize_inventory_path(

                self.interop_widgets["InventoryExcelPath"].text().strip()

            )

            self.interop_widgets["InventoryExcelPath"].setText(inventory_path)

            QSettings().setValue("SOI_MetadataGenerator/fixed_inventory_path", inventory_path)



            created = False

            if not os.path.exists(inventory_path):

                write_inventory_xlsx(inventory_path, [])

                created = True

            else:

                # Validate that the workbook is readable. If a zero-byte/corrupt file

                # exists at the path, recreate a clean inventory with headers.

                try:

                    read_inventory_xlsx(inventory_path)

                except Exception:

                    write_inventory_xlsx(inventory_path, [])

                    created = True



            if created:

                self.set_status(f"Inventory created: {inventory_path}", True)

                if show_message:

                    QMessageBox.information(

                        self,

                        "Inventory Created",

                        f"Metadata Inventory Excel has been created:\n\n{inventory_path}"

                    )

            else:

                if show_message:

                    self.set_status(f"Inventory available: {inventory_path}", True)

                    QMessageBox.information(

                        self,

                        "Inventory Available",

                        f"Metadata Inventory Excel is available:\n\n{inventory_path}"

                    )

            return inventory_path

        except Exception as exc:

            self.set_status(f"Inventory creation failed: {exc}", False)

            if show_message:

                QMessageBox.warning(self, "Inventory Error", str(exc))

            return ""



    def default_xml_folder(self):

        """Return the fixed folder where generated XML files should be kept.

        The inventory register path stays fixed, and all default XML files are kept

        beside that register unless the user manually chooses another XML path.

        """

        inventory_path = self.normalize_inventory_path(

            self.interop_widgets["InventoryExcelPath"].text().strip()

        )

        folder = os.path.dirname(inventory_path) or os.path.expanduser("~")

        if folder and not os.path.exists(folder):

            os.makedirs(folder, exist_ok=True)

        return folder



    def safe_dataset_file_stem(self, layer=None, source_path=""):

        """Create a stable safe file/layer name for default XML naming."""

        name = ""

        if layer is not None:

            try:

                name = layer.name()

            except Exception:

                name = ""

        if not name:

            fs = self.clean_source_for_filesystem(source_path)

            if fs:

                name = os.path.splitext(os.path.basename(fs.rstrip(os.sep)))[0]

        if not name:

            name = self.get_field_by_label("File name") or "SOI_ISO19115"

        name = re.sub(r"[^A-Za-z0-9_\-]+", "_", safe_text(name)).strip("_")

        return name or "SOI_ISO19115"



    def default_xml_path_for_current_selection(self, layer=None, source_path=""):

        """Default XML path: fixed XML folder + selected file/layer name."""

        stem = self.safe_dataset_file_stem(layer=layer, source_path=source_path)

        return os.path.join(self.default_xml_folder(), stem + "_SOI_ISO19115_metadata.xml")



    def set_default_xml_path_for_selection(self, layer=None, source_path="", force=True):

        """Set XML path for current layer.

        force=True is used on layer change so a new layer gets its own default XML.

        Registered datasets will overwrite this with the XML path from inventory.

        """

        xml_path = self.default_xml_path_for_current_selection(layer=layer, source_path=source_path)

        current = safe_text(self.interop_widgets["MetadataXMLPath"].text())

        if force or not current:

            self.interop_widgets["MetadataXMLPath"].setText(xml_path)

        return xml_path



    def extract_and_fill_silent(self):

        """Extract auto-derived metadata without message boxes. Used on layer change."""

        return self.extract_and_fill(show_message=False)



    def set_status(self, text, ok=None):

        """Show one of the four registry states, with details in the tooltip."""

        if not hasattr(self, "registry_status_label"):

            return

        lowered = safe_text(text).lower()

        if "error" in lowered or "failed" in lowered:

            label, color = "🔴 Inventory Error", "#c62828"

        elif "not registered" in lowered or "not registered yet" in lowered:

            label, color = "🔵 New Dataset", "#1565c0"

        elif "registered" in lowered and ("missing" in lowered or "not exist" in lowered):

            label, color = "🟡 Registered - XML Missing", "#f9a825"

        elif "registered" in lowered:

            label, color = "🟢 Registered", "#2e7d32"

        elif "no dataset" in lowered or "inventory available" in lowered or "inventory created" in lowered:

            label, color = "—", "#546e7a"

        else:

            label, color = "🔵 New Dataset", "#1565c0"

        self.registry_status_label.setText("Registry Status: " + label)

        self.registry_status_label.setToolTip(safe_text(text))

        self.registry_status_label.setStyleSheet(f"font-weight:bold; color:{color}; padding:4px;")


    def on_dataset_selection_changed(self, *args):

        """When the selected layer changes, update paths and auto-load registered metadata if present.

        Startup and placeholder selections are ignored so no alert appears when the tool opens.

        """

        if getattr(self, "_initializing", False):

            return



        mode = self.source_type_combo.currentIndex() if hasattr(self, "source_type_combo") else 0

        if mode == 0 and self.project_layer_combo.currentData() is None:

            self.interop_widgets["DatasetStoragePath"].clear()

            self.interop_widgets["QGISLayerName"].clear()

            self.interop_widgets["MetadataXMLPath"].clear()

            self.set_status("No dataset selected", None)

            return

        if mode in (2, 3) and self.container_layer_combo.currentData() is None:

            self.set_status("Container selected. Choose a layer/raster to check registry.", None)

            return



        try:

            layer, source_path = self.current_selected_layer_or_path()

        except Exception:

            return



        if layer is not None and layer.isValid():

            try:

                self.interop_widgets["QGISLayerName"].setText(layer.name())

            except Exception:

                pass

            if not source_path:

                try:

                    source_path = layer.source()

                except Exception:

                    source_path = ""



        if source_path:

            self.interop_widgets["DatasetStoragePath"].setText(source_path)

            # Keep inventory path fixed. Do not change it when the selected layer/dataset changes.

            self.ensure_inventory_path_exists_in_ui()



            # Each selected layer gets its own default XML path in the fixed register folder.

            # If the layer is already registered, the inventory record will overwrite this

            # with the existing linked XML path.

            self.set_default_xml_path_for_selection(layer=layer, source_path=source_path, force=True)



            # Auto-extract factual metadata immediately after layer/dataset change.

            self.extract_and_fill_silent()



            # Finally check the fixed inventory register and auto-load linked XML if present.

            self.check_existing_inventory_for_selected_dataset(show_no_record=False, silent=True)

        else:

            self.set_status("No dataset selected", None)



    def browse_inventory_excel(self):

        file_path, _ = QFileDialog.getSaveFileName(

            self,

            "Select Metadata Inventory Excel",

            self.interop_widgets["InventoryExcelPath"].text() or "metadata_inventory_register.xlsx",

            "Excel Files (*.xlsx)"

        )

        if file_path:

            file_path = self.normalize_inventory_path(file_path)

            self.interop_widgets["InventoryExcelPath"].setText(file_path)

            QSettings().setValue("SOI_MetadataGenerator/fixed_inventory_path", file_path)

            self.ensure_inventory_file_exists(show_message=True)

            try:

                layer, source_path = self.current_selected_layer_or_path()

                if layer is not None and layer.isValid():

                    self.set_default_xml_path_for_selection(layer=layer, source_path=source_path, force=True)

            except Exception:

                pass





    def browse_existing_xml(self):

        xml_path, _ = QFileDialog.getOpenFileName(

            self,

            "Select Existing Metadata XML",

            os.path.dirname(self.raster_path_edit.text().strip()) if self.raster_path_edit.text().strip() else "",

            "XML Files (*.xml);;All Files (*.*)"

        )

        if xml_path:

            self.interop_widgets["MetadataXMLPath"].setText(xml_path)



    def load_xml_into_gui(self, show_messages=True):

        xml_path = self.interop_widgets["MetadataXMLPath"].text().strip()

        if not xml_path or not os.path.exists(xml_path):

            QMessageBox.warning(self, "Missing XML", "Please select an existing metadata XML file first.")

            return False



        try:

            xml_values, xml_interop, xml_inventory = read_metadata_xml_values(xml_path)

        except Exception as exc:

            QMessageBox.critical(self, "XML Read Error", f"Could not read metadata XML:\n\n{exc}")

            return False



        if not xml_values and not xml_interop and not xml_inventory:

            QMessageBox.warning(

                self,

                "Unsupported XML",

                "The selected XML was linked, but no matching GUI fields could be read from it.\n\n"

                "This loader can fully restore XML generated by this tool. Other SOI/ISO XML files can still be added to the inventory."

            )

            return False



        label_map = {f["id"]: f["label"] for f in SOI_TEMPLATE_FIELDS}

        current_values = {f["id"]: self.get_widget_value(self.widgets[f["id"]]) for f in SOI_TEMPLATE_FIELDS if f["id"] in self.widgets}

        changes = compare_records(current_values, xml_values, label_map=label_map, ignore_blank_old=True)



        for fid, value in xml_values.items():

            widget = self.widgets.get(fid)

            if widget is not None:

                if isinstance(widget, QTextEdit):

                    widget.setPlainText(safe_text(value))

                elif isinstance(widget, QComboBox):

                    idx = widget.findText(safe_text(value))

                    if idx < 0 and safe_text(value):

                        widget.addItem(safe_text(value))

                        idx = widget.findText(safe_text(value))

                    if idx >= 0:

                        widget.setCurrentIndex(idx)

                else:

                    widget.setText(safe_text(value))



        for key, value in xml_interop.items():
            if key == "InventoryCSVPath":
                key = "InventoryExcelPath"
            if key in self.interop_widgets and safe_text(value):
                self.interop_widgets[key].setText(safe_text(value))


        for key, value in xml_inventory.items():
            self.set_soi_inventory_field(key, value, overwrite=True)

        self.interop_widgets["MetadataXMLPath"].setText(xml_path)

        self.apply_readonly_rules()



        if show_messages:

            if changes:

                msg = "XML loaded. The following existing GUI values were different from the XML:\n\n"

                for label, old, new in changes[:20]:

                    msg += f"- {label}:\n  GUI: {old}\n  XML: {new}\n"

                if len(changes) > 20:

                    msg += f"\n... and {len(changes) - 20} more changes."

                QMessageBox.warning(self, "XML Loaded - Differences Found", msg)

            else:

                QMessageBox.information(self, "XML Loaded", "Metadata XML loaded into the GUI successfully.")

        return True



    def build_inventory_row_from_gui(self, xml_path=""):
        """Build only the reduced SOI columns plus technical match columns."""
        layer, selected_source = self.current_selected_layer_or_path()
        data_path = selected_source or self.interop_widgets["DatasetStoragePath"].text().strip()
        layer_name = layer.name() if layer is not None and layer.isValid() else self.interop_widgets["QGISLayerName"].text().strip()
        self.populate_soi_inventory_defaults(
            metadata=getattr(self, "current_raster_metadata", {}),
            layer=layer,
            source_path=data_path,
            overwrite=False,
        )

        row = {header: self.get_soi_inventory_value(header) for header in SOI_INVENTORY_HEADERS if header != "SlNO"}
        dataset_id = self.get_field_by_label("metadataIdentifier · code")
        key = canonical_dataset_key(data_path, layer_name)
        row.update({
            "Dataset_Key": key,
            "Dataset_Hash": hashlib.sha1(key.encode("utf-8")).hexdigest() if key else "",
            "Dataset_ID": dataset_id,
            "Data_Path": data_path,
            "Metadata_XML_Path": xml_path or self.interop_widgets["MetadataXMLPath"].text().strip(),
            "Layer_Name": layer_name,
            "Dataset_Modified_Time": file_modified_iso(self.clean_source_for_filesystem(data_path)),
            "Metadata_XML_Modified_Time": file_modified_iso(xml_path),
            "Last_Modified": now_iso(),
        })
        return row

    def add_or_update_inventory_from_existing_xml(self):

        xml_path = self.interop_widgets["MetadataXMLPath"].text().strip()

        if not xml_path:

            self.browse_existing_xml()

            xml_path = self.interop_widgets["MetadataXMLPath"].text().strip()

        if not xml_path or not os.path.exists(xml_path):

            QMessageBox.warning(self, "Missing XML", "Please select a valid metadata XML file.")

            return



        inventory_path = self.normalize_inventory_path(self.interop_widgets["InventoryExcelPath"].text().strip())

        self.interop_widgets["InventoryExcelPath"].setText(inventory_path)

        QSettings().setValue("SOI_MetadataGenerator/fixed_inventory_path", inventory_path)



        # Try loading XML values first; if it is not this tool's XML, still add/update the inventory with the selected paths.

        if os.path.exists(xml_path):

            try:

                self.load_xml_into_gui(show_messages=False)

            except Exception:

                pass



        row = self.build_inventory_row_from_gui(xml_path=xml_path)

        try:

            action, inventory_path, excel_row, total_records = upsert_inventory_xlsx(inventory_path, row)

            self.interop_widgets["InventoryExcelPath"].setText(inventory_path)

            QMessageBox.information(

                self,

                "Inventory Updated",

                f"Inventory {action}.\n\nWorkbook:\n{inventory_path}\n\nExcel Row: {excel_row}\nTotal Records: {total_records}"

            )

        except Exception as exc:

            QMessageBox.critical(self, "Inventory Error", f"Could not update inventory Excel:\n\n{exc}")



    def check_existing_inventory_for_selected_dataset(self, show_no_record=False, silent=False):

        inventory_path = self.interop_widgets["InventoryExcelPath"].text().strip()

        try:

            _layer, selected_source = self.current_selected_layer_or_path()

        except Exception:

            selected_source = ""

        data_path = selected_source or self.interop_widgets["DatasetStoragePath"].text().strip() or self.raster_path_edit.text().strip()

        if inventory_path:

            inventory_path = self.ensure_inventory_file_exists(show_message=False)

        if not inventory_path or not data_path:

            self.set_status("Not registered yet" if data_path else "No dataset selected", False if data_path else None)

            if show_no_record and not silent:

                QMessageBox.information(self, "No Inventory Record", "Dataset path is not available yet.")

            return None

        try:

            records = read_inventory_xlsx(inventory_path)

        except Exception as exc:

            self.set_status(f"Inventory read error: {exc}", False)

            if not silent:

                QMessageBox.warning(self, "Inventory Read Error", str(exc))

            return None

        try:
            layer_for_key, _src_for_key = self.current_selected_layer_or_path()
            layer_name_for_key = layer_for_key.name() if layer_for_key is not None else self.interop_widgets["QGISLayerName"].text()
        except Exception:
            layer_name_for_key = self.interop_widgets["QGISLayerName"].text()
        idx, record = find_inventory_record(
            records,
            data_path=data_path,
            dataset_hash_value=dataset_hash(data_path, layer_name_for_key),
            layer_name=layer_name_for_key
        )

        if record is None:

            self.set_status("Not registered in selected inventory; default XML path prepared", False)

            if show_no_record and not silent:

                QMessageBox.information(self, "No Inventory Record", "No matching inventory record found for this dataset path.")

            return None



        xml_path = record.get("XML_Path") or record.get("Metadata_XML_Path", "")

        if xml_path:

            self.interop_widgets["MetadataXMLPath"].setText(xml_path)

        if xml_path and os.path.exists(xml_path):

            self.load_xml_into_gui(show_messages=False)

        else:

            self.set_status("Registered - XML Missing", False)

            return record



        change_lines = []

        checks = {

            "Dataset file modified time": (record.get("Dataset_Modified_Time", ""), file_modified_iso(self.clean_source_for_filesystem(data_path))),

            "Dataset file size MB": (record.get("Dataset_File_Size_MB", ""), file_size_mb(self.clean_source_for_filesystem(data_path))),

            "Metadata XML modified time": (record.get("Metadata_XML_Modified_Time", ""), file_modified_iso(xml_path)),

        }

        for label, (old, new) in checks.items():

            if safe_text(old) and safe_text(new) and safe_text(old) != safe_text(new):

                change_lines.append(f"- {label}: inventory={old}, current={new}")



        if change_lines:

            self.set_status("Registered; linked XML loaded; file-level changes detected", False)

            if not silent:

                QMessageBox.warning(

                    self,

                    "Existing Metadata Found - Changes Detected",

                    "An inventory record and linked XML were found for this dataset.\n\n"

                    "The XML has been loaded. The following file-level changes were detected:\n\n" + "\n".join(change_lines)

                )

        else:

            self.set_status("Registered; linked XML loaded", True)

            if not silent:

                QMessageBox.information(

                    self,

                    "Existing Metadata Found",

                    "An inventory record was found and the linked XML was loaded. No file-level changes were detected."

                )

        return record



    def check_inventory_xml_changes(self):

        record = self.check_existing_inventory_for_selected_dataset(show_no_record=True)

        if not record:

            return



        xml_path = self.interop_widgets["MetadataXMLPath"].text().strip()

        if not xml_path or not os.path.exists(xml_path):

            QMessageBox.warning(self, "Missing Linked XML", "Inventory record exists, but the linked XML file does not exist at the stored path.")

            return



        try:

            xml_values, _, _ = read_metadata_xml_values(xml_path)

        except Exception as exc:

            QMessageBox.critical(self, "XML Read Error", str(exc))

            return



        label_map = {f["id"]: f["label"] for f in SOI_TEMPLATE_FIELDS}

        current_values = {f["id"]: self.get_widget_value(self.widgets[f["id"]]) for f in SOI_TEMPLATE_FIELDS if f["id"] in self.widgets}

        changes = compare_records(xml_values, current_values, label_map=label_map)



        msg = "Inventory and XML check completed."

        if changes:

            msg += "\n\nGUI values differ from linked XML:\n\n"

            for label, old, new in changes[:25]:

                msg += f"- {label}:\n  XML: {old}\n  GUI: {new}\n"

            if len(changes) > 25:

                msg += f"\n... and {len(changes) - 25} more changed fields."

            QMessageBox.warning(self, "Changes Detected", msg)

        else:

            QMessageBox.information(self, "No Metadata Field Changes", msg + "\n\nNo GUI/XML field differences were found.")



    def set_field_by_label(self, label, value):

        fid = self.field_lookup.get(label)

        if not fid:

            return

        widget = self.widgets.get(fid)

        if widget is None:

            return



        if isinstance(widget, QTextEdit):

            widget.setPlainText(safe_text(value))

        elif isinstance(widget, QComboBox):

            val = safe_text(value)

            idx = widget.findText(val)

            if idx < 0 and val:

                widget.addItem(val)

                idx = widget.findText(val)

            if idx >= 0:

                widget.setCurrentIndex(idx)

        else:

            widget.setText(safe_text(value))



    def get_widget_value(self, widget):

        if isinstance(widget, QTextEdit):

            return widget.toPlainText().strip()

        if isinstance(widget, QComboBox):

            return widget.currentText().strip()

        return widget.text().strip()



    def get_field_by_label(self, label):

        fid = self.field_lookup.get(label)

        if not fid:

            return ""

        widget = self.widgets.get(fid)

        if not widget:

            return ""

        return self.get_widget_value(widget)



    def apply_readonly_rules(self):

        unlock = self.unlock_checkbox.isChecked()



        for field in SOI_TEMPLATE_FIELDS:

            widget = self.widgets.get(field["id"])

            if not widget:

                continue



            is_auto = field["label"] in AUTO_DERIVED_LABELS



            if isinstance(widget, QTextEdit):

                widget.setReadOnly(is_auto and not unlock)

            elif isinstance(widget, QComboBox):

                widget.setEnabled((not is_auto) or unlock)

            else:

                widget.setReadOnly(is_auto and not unlock)



            if is_auto and not unlock:

                if isinstance(widget, QComboBox):

                    widget.setStyleSheet("""

                        QComboBox { background-color:#eeeeee; color:#555555; border:1px solid #b0bec5; border-radius:4px; padding:5px; padding-right:24px; }

                        QComboBox QLineEdit { background-color:#eeeeee; color:#555555; border:0px; padding:0px 2px; margin:0px; }

                        QComboBox::drop-down { width:22px; border-left:1px solid #b0bec5; background-color:#e0e0e0; }

                        QComboBox QAbstractItemView { background-color:#ffffff; color:#111111; selection-background-color:#e3f2fd; selection-color:#000000; }

                    """)

                    if widget.lineEdit():

                        widget.lineEdit().setStyleSheet("background-color:#eeeeee; color:#555555; border:0px; padding:0px 2px; margin:0px;")

                else:

                    widget.setStyleSheet("background-color:#eeeeee; color:#555555; border:1px solid #b0bec5; border-radius:4px; padding:5px;")

            else:

                if isinstance(widget, QComboBox):

                    widget.setStyleSheet("""

                        QComboBox { background-color:#ffffff; color:#111111; border:1px solid #b0bec5; border-radius:4px; padding:5px; padding-right:24px; }

                        QComboBox QLineEdit { background-color:#ffffff; color:#111111; border:0px; padding:0px 2px; margin:0px; }

                        QComboBox::drop-down { width:22px; border-left:1px solid #b0bec5; background-color:#eef5f9; }

                        QComboBox QAbstractItemView { background-color:#ffffff; color:#111111; selection-background-color:#e3f2fd; selection-color:#000000; }

                    """)

                    if widget.lineEdit():

                        widget.lineEdit().setStyleSheet("background-color:#ffffff; color:#111111; border:0px; padding:0px 2px; margin:0px;")

                else:

                    widget.setStyleSheet("background-color:#ffffff; color:#111111; border:1px solid #b0bec5; border-radius:4px; padding:5px;")




        for header, widget in getattr(self, "soi_inventory_widgets", {}).items():
            is_auto = header in AUTO_DERIVED_INVENTORY_HEADERS
            read_only = is_auto and not unlock
            if isinstance(widget, QTextEdit):
                widget.setReadOnly(read_only)
            else:
                widget.setReadOnly(read_only)
            widget.setStyleSheet(
                "background-color:#eeeeee; color:#555555; border:1px solid #b0bec5; border-radius:4px; padding:5px;"
                if read_only else
                "background-color:#ffffff; color:#111111; border:1px solid #b0bec5; border-radius:4px; padding:5px;"
            )

    def extract_and_fill(self, show_message=True):

        layer, source_path = self.current_selected_layer_or_path()



        if layer is None or not layer.isValid():

            if show_message:

                QMessageBox.warning(

                    self,

                    "Missing or Invalid Dataset",

                    "Please select a valid project layer, file dataset, or .gdb/.gpkg vector/raster item."

                )

            return False



        metadata = extract_layer_metadata(layer, source_path=source_path)

        self.current_raster_metadata = metadata



        if not metadata.get("valid"):

            if show_message:

                QMessageBox.critical(self, "Invalid Dataset", "QGIS could not read metadata from this layer.")

            return False



        base = metadata.get("File name") or layer.name()



        # Dataset file information directly available from selected layer/file.

        self.set_field_by_label("File name", base)



        # Spatial representation.

        if metadata.get("layer_type") == "Raster":

            self.set_field_by_label("spatialRepresentationType", "grid")

            self.set_field_by_label("spatialResolution · distance (GSD)", metadata["pixel_width"])

            self.set_field_by_label("numberOfDimensions", metadata["number_of_dimensions"])

            self.set_field_by_label("axisDimensionProperties · column · dimensionSize", metadata["width"])

            self.set_field_by_label("axisDimensionProperties · row · dimensionSize", metadata["height"])

            self.set_field_by_label("cellGeometry", metadata["cell_geometry"])

            if not self.get_field_by_label("attributeDescription") and metadata.get("bands") != "":

                self.set_field_by_label("attributeDescription", f"Raster band count: {metadata['bands']}")

        elif metadata.get("layer_type") == "Vector":

            self.set_field_by_label("spatialRepresentationType", "vector")

            if not self.get_field_by_label("geometricObjects · geometricObjectCount") and metadata.get("feature_count") != "":

                self.set_field_by_label("geometricObjects · geometricObjectCount", metadata["feature_count"])



        # Bounding box directly derived from layer extent.

        self.set_field_by_label("geographicElement · westBoundLongitude", metadata["west"])

        self.set_field_by_label("geographicElement · eastBoundLongitude", metadata["east"])

        self.set_field_by_label("geographicElement · southBoundLatitude", metadata["south"])

        self.set_field_by_label("geographicElement · northBoundLatitude", metadata["north"])



        # CRS directly available from layer.

        self.set_field_by_label("Horizontal CRS · code", metadata["crs_authid"])

        self.set_field_by_label("CRS description / notes", metadata["crs_description"])



        # Format/provider directly available.

        if metadata.get("format_name"):

            self.set_field_by_label("distributionFormat · name", metadata["format_name"])



        # Interoperability register values.

        dataset_source = metadata.get("file_path") or source_path

        self.interop_widgets["DatasetStoragePath"].setText(dataset_source)

        self.interop_widgets["QGISLayerName"].setText(layer.name())

        self.ensure_inventory_path_exists_in_ui()

        self.set_default_xml_path_for_selection(layer=layer, source_path=dataset_source, force=False)



        # Pre-fill SOI inventory tab from extracted dataset properties.

        self.populate_soi_inventory_defaults(metadata=metadata, layer=layer, source_path=dataset_source)



        self.apply_readonly_rules()

        if show_message:

            QMessageBox.information(self, "Done", "Dataset metadata extracted and SOI ISO fields filled.")

        return True



    def fill_sample_values(self):

        """Deprecated: intentionally does nothing. Metadata should not be guessed."""

        QMessageBox.information(

            self,

            "No Sample Values",

            "Sample/assumed metadata is disabled. Blank metadata fields must be filled by the user."

        )



    def validate_required_fields(self):

        missing = []

        for field in SOI_TEMPLATE_FIELDS:

            if field["required"].lower() == "yes":

                widget = self.widgets.get(field["id"])

                if widget and not self.get_widget_value(widget):

                    missing.append(f'{field["section"]} > {field["label"]}')

        return missing



    def collect_values(self):

        values = {}

        for field in SOI_TEMPLATE_FIELDS:

            widget = self.widgets.get(field["id"])

            values[field["id"]] = self.get_widget_value(widget) if widget else ""



        values["_interop"] = {

            key: self.get_widget_value(widget)

            for key, widget in self.interop_widgets.items()

        }

        values["_soi_inventory"] = {

            key: self.get_widget_value(widget)

            for key, widget in getattr(self, "soi_inventory_widgets", {}).items()

        }

        return values



    def generate_xml_and_update_inventory(self, checked=False):
        """Qt slot for the combined action; ignore QPushButton's checked value."""
        del checked
        self.generate_xml_and_inventory(update_inventory=True)

    def generate_xml_only(self, checked=False):
        """Qt slot for XML-only generation; ignore QPushButton's checked value."""
        del checked
        self.generate_xml_and_inventory(update_inventory=False)

    def generate_xml_and_inventory(self, update_inventory=True):

        missing = self.validate_required_fields()

        if missing:

            preview = "\n".join(missing[:12])

            extra = "" if len(missing) <= 12 else f"\n... and {len(missing) - 12} more"

            reply = QMessageBox.question(

                self,

                "Required Fields Missing",

                f"The following required fields are empty:\n\n{preview}{extra}\n\nGenerate XML anyway?",

                QMessageBox.Yes | QMessageBox.No

            )

            if reply != QMessageBox.Yes:

                return



        layer, source_path = self.current_selected_layer_or_path()

        source_path = source_path or self.interop_widgets["DatasetStoragePath"].text().strip() or self.raster_path_edit.text().strip()

        if layer is not None and layer.isValid():

            self.interop_widgets["QGISLayerName"].setText(layer.name())

        if source_path:

            self.interop_widgets["DatasetStoragePath"].setText(source_path)



        # Use the per-layer default XML path in the fixed register/XML folder.

        # This path is also refreshed whenever the selected layer changes.

        default_xml_path = safe_text(self.interop_widgets["MetadataXMLPath"].text())

        if not default_xml_path:

            default_xml_path = self.set_default_xml_path_for_selection(layer=layer, source_path=source_path, force=True)



        output_xml_path, _ = QFileDialog.getSaveFileName(

            self,

            "Save SOI Metadata XML",

            default_xml_path,

            "XML Files (*.xml)"

        )



        if not output_xml_path:

            return



        if not output_xml_path.lower().endswith(".xml"):

            output_xml_path += ".xml"



        self.populate_soi_inventory_defaults(
            metadata=getattr(self, "current_raster_metadata", {}),
            layer=layer,
            source_path=source_path,
            overwrite=False,
        )
        values = self.collect_values()

        values["_interop"]["DatasetStoragePath"] = source_path

        values["_interop"]["MetadataXMLPath"] = output_xml_path

        self.interop_widgets["MetadataXMLPath"].setText(output_xml_path)



        try:

            write_metadata_xml(values, output_xml_path)

        except Exception as exc:

            QMessageBox.critical(self, "XML Error", f"Failed to generate XML:\n\n{exc}")

            return



        inventory_msg = ""

        if update_inventory:

            inventory_path = self.ensure_inventory_file_exists(show_message=False)

            if not inventory_path:

                inventory_path = self.normalize_inventory_path(self.interop_widgets["InventoryExcelPath"].text().strip())

            self.interop_widgets["InventoryExcelPath"].setText(inventory_path)

            QSettings().setValue("SOI_MetadataGenerator/fixed_inventory_path", inventory_path)



            # Ensure the metadata identifier is stable and also written into the UI for future matching.

            dataset_id = self.get_field_by_label("metadataIdentifier · code")

            if not dataset_id:

                dataset_id = str(uuid.uuid4())

                self.set_field_by_label("metadataIdentifier · code", dataset_id)

                values = self.collect_values()

                values["_interop"]["DatasetStoragePath"] = source_path

                values["_interop"]["MetadataXMLPath"] = output_xml_path

                # Re-write XML so generated UUID is also stored inside metadata.

                write_metadata_xml(values, output_xml_path)



            row = self.build_inventory_row_from_gui(xml_path=output_xml_path)


            row["Dataset_ID"] = dataset_id



            row["Data_Path"] = source_path


            row["Metadata_XML_Path"] = output_xml_path

            row["Last_Modified"] = now_iso()


            row["Dataset_Modified_Time"] = file_modified_iso(self.clean_source_for_filesystem(source_path))

            row["Metadata_XML_Modified_Time"] = file_modified_iso(output_xml_path)



            try:

                action, inventory_path, excel_row, total_records = upsert_inventory_xlsx(inventory_path, row)

                self.interop_widgets["InventoryExcelPath"].setText(inventory_path)

                self.set_status(f"Inventory {action}; Excel row {excel_row}; metadata registered", True)

                inventory_msg = f"\n\nInventory {action}:\n{inventory_path}\n\nExcel Row: {excel_row}\nTotal Records: {total_records}"

            except Exception as exc:

                self.set_status(f"Inventory update failed: {exc}", False)

                inventory_msg = f"\n\nXML generated, but inventory update failed:\n{exc}"



        QMessageBox.information(

            self,

            "Completed",

            f"Metadata XML generated successfully:\n{output_xml_path}{inventory_msg}"

        )





dialog = SOIMetadataDialog()

dialog.show()

